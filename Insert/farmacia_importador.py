"""
IMPORTADOR DE DATOS PARA FARMACIA
Versión con interfaz gráfica - Todo en un solo archivo
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import re
from datetime import datetime, date
from calendar import monthrange
import os
import pyodbc
import csv
import threading
import sys
from tkinter import font as tkfont


# ==================== CLASE EXTRACTORA (del script original) ====================
class ExtraerDatosFarmacia:
    def __init__(self, archivo_excel):
        self.archivo_excel = archivo_excel
        self.medicamentos = []
        self.inventario = []
        self.movimientos = []
        self.wb = None
        self.col_mapping = {}
        self.usuario_admin_nombre = None
        self.usuario_admin_id = None
        self.log_file = None
        self.omitidos = []
        self.log_callback = None  # Para enviar logs a la GUI

    def _log(self, mensaje, nivel="INFO"):
        """Escribe en el archivo de log si está abierto, y también en consola y GUI"""
        if self.log_file:
            self.log_file.write(f"[{nivel}] {mensaje}\n")
            self.log_file.flush()
        print(mensaje)
        if self.log_callback:
            self.log_callback(mensaje)

    def _obtener_usuario_admin(self, cursor):
        """Obtiene el usuario administrador"""
        try:
            cursor.execute("""
                SELECT id_usuario, nombre_usuario
                FROM usuarios
                WHERE nombre_usuario = 'admin'
            """)
            row = cursor.fetchone()
            if row:
                self.usuario_admin_id = row[0]
                self.usuario_admin_nombre = row[1]
                self._log(f"  👤 Usuario admin encontrado: {self.usuario_admin_nombre} (ID: {self.usuario_admin_id})")
            else:
                cursor.execute("SELECT TOP 1 id_usuario, nombre_usuario FROM usuarios ORDER BY id_usuario")
                row = cursor.fetchone()
                if row:
                    self.usuario_admin_id = row[0]
                    self.usuario_admin_nombre = row[1]
                    self._log(f"  👤 Usuario por defecto: {self.usuario_admin_nombre} (ID: {self.usuario_admin_id})")
                else:
                    self.usuario_admin_nombre = "admin"
                    self.usuario_admin_id = None
                    self._log(f"  ⚠️ No se encontraron usuarios, se usará 'admin'")
            return True
        except Exception as e:
            self._log(f"  ⚠️ Error al obtener usuario admin: {e}")
            self.usuario_admin_nombre = "admin"
            self.usuario_admin_id = None
            return True

    def _parsear_fecha_caducidad(self, valor):
        """Parsea fecha de caducidad"""
        if pd.isna(valor):
            return None
        try:
            if isinstance(valor, datetime):
                return valor.date()
            elif isinstance(valor, str):
                valor = valor.strip()
                if valor.lower() in ('nan', 'none', ''):
                    return None
                for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y']:
                    try:
                        return datetime.strptime(valor, fmt).date()
                    except:
                        continue
                patron_mes_anio = r'(\d{1,2})[/-](\d{4})'
                match = re.search(patron_mes_anio, valor)
                if match:
                    mes = int(match.group(1))
                    anio = int(match.group(2))
                    if mes > 12:
                        mes, anio = anio, mes
                    ultimo_dia = monthrange(anio, mes)[1]
                    return date(anio, mes, ultimo_dia)
                patron_anio = r'(\d{4})'
                match = re.search(patron_anio, valor)
                if match:
                    anio = int(match.group(1))
                    return date(anio, 12, 31)
            elif hasattr(valor, 'date'):
                return valor.date()
        except Exception as e:
            self._log(f"  ⚠️ Error al parsear fecha: {valor} - {e}")
        return None

    def _extraer_numero(self, valor):
        """Extrae un número entero"""
        if pd.isna(valor):
            return 0
        try:
            if isinstance(valor, (int, float)):
                return int(valor)
            elif isinstance(valor, str):
                numeros = re.findall(r'\d+', valor)
                if numeros:
                    return sum(int(n) for n in numeros)
        except:
            pass
        return 0

    def _detectar_color_morado(self, ws, fila, columna_letra='A'):
        """Detecta color morado de fondo"""
        try:
            celda = ws[f"{columna_letra}{fila}"]
            if celda.fill and hasattr(celda.fill, 'fgColor') and celda.fill.fgColor:
                fg_color = celda.fill.fgColor
                if hasattr(fg_color, 'rgb') and fg_color.rgb:
                    rgb = str(fg_color.rgb).upper()
                    if len(rgb) >= 6:
                        if len(rgb) == 8:
                            rgb = rgb[2:]
                        try:
                            r = int(rgb[0:2], 16)
                            g = int(rgb[2:4], 16)
                            b = int(rgb[4:6], 16)
                            if (r > 100 and b > 100 and g < 150) or (r > 150 and b > 150):
                                return True
                        except:
                            pass
        except:
            pass
        return False

    def _extraer_concentracion(self, nombre, presentacion):
        """Extrae la concentración del medicamento"""
        texto = f"{nombre} {presentacion}".lower()
        patrones = [
            r'(\d+(?:\.\d+)?\s*(?:mg|mcg|µg|g|ml|UI)(?:\s*\/\s*(?:ml|mg|g))?)',
            r'(\d+(?:\.\d+)?\s*mg(?:\s*\/\s*\d+(?:\.\d+)?\s*ml)?)',
            r'(\d+(?:\.\d+)?\s*(?:%|por ciento))',
        ]
        for patron in patrones:
            coincidencia = re.search(patron, texto, re.IGNORECASE)
            if coincidencia:
                return coincidencia.group(1)
        return None

    def _extraer_forma_farmaceutica(self, presentacion, nombre):
        """Extrae la forma farmacéutica"""
        texto = f"{presentacion} {nombre}".lower()
        formas = {
            'Tableta': ['tableta', 'tabletas', 'comprimido', 'comprimidos', 'tab', 'tabs'],
            'Cápsula': ['cápsula', 'capsula', 'cápsulas', 'capsulas', 'cap'],
            'Jarabe': ['jarabe', 'solución oral', 'suspensión oral', 'suspension oral'],
            'Inyectable': ['inyectable', 'ampolleta', 'ampolla', 'vial', 'jeringa'],
            'Crema': ['crema', 'ungüento', 'pomada', 'gel'],
            'Polvo': ['polvo', 'granulado', 'sobres'],
            'Gotas': ['gotas', 'gotero'],
            'Aerosol': ['aerosol', 'inhalador', 'spray'],
            'Solución': ['solución', 'solucion', 'solución inyectable'],
        }
        for forma, palabras in formas.items():
            for palabra in palabras:
                if palabra in texto:
                    return forma
        return None

    def _limpiar_nombre_medicamento(self, nombre):
        """Limpia el nombre eliminando concentración y forma farmacéutica"""
        nombre_limpio = re.sub(r'\s+\d+(?:\.\d+)?\s*(?:mg|mcg|µg|g|ml|UI).*$', '', nombre, flags=re.IGNORECASE)
        nombre_limpio = re.sub(r'\s*\(.*?\)\s*$', '', nombre_limpio)
        palabras_eliminar = ['tableta', 'tabletas', 'cápsula', 'capsula', 'comprimido', 'comprimidos',
                             'gragea', 'grageas', 'sobre', 'sobres', 'vial', 'ampolla']
        for palabra in palabras_eliminar:
            if nombre_limpio.lower().endswith(palabra):
                nombre_limpio = nombre_limpio[:-len(palabra)].strip()
                break
        if nombre_limpio and len(nombre_limpio) > 1:
            nombre_limpio = nombre_limpio[0].upper() + nombre_limpio[1:].lower()
        elif nombre_limpio:
            nombre_limpio = nombre_limpio.upper()
        return nombre_limpio.strip()

    def _eliminar_duplicados_catalogo(self):
        """Elimina medicamentos duplicados del catálogo"""
        medicamentos_unicos = []
        claves_vistas = set()
        for med in self.medicamentos:
            clave = f"{med['nombre'].lower()}|{med.get('forma_farmaceutica', '')}|{med.get('concentracion', '')}"
            if clave not in claves_vistas:
                claves_vistas.add(clave)
                medicamentos_unicos.append(med)
        self.medicamentos = medicamentos_unicos

    def _mapear_columnas(self, columnas):
        """Mapea los nombres de columnas"""
        self.col_mapping = {}
        for i, col in enumerate(columnas):
            col_upper = col.upper() if col else ''
            if 'MEDICAMENTO' in col_upper:
                self.col_mapping['medicamento'] = col
            elif 'PRESENTACIÓN' in col_upper or 'PRESENTACION' in col_upper:
                self.col_mapping['presentacion'] = col
            elif 'PROVEEDOR' in col_upper:
                self.col_mapping['proveedor'] = col
            elif 'ENTRADA' in col_upper or 'ENTRADAS' in col_upper:
                self.col_mapping['entradas'] = col
            elif 'SALIDA' in col_upper or 'SALIDAS' in col_upper:
                self.col_mapping['salidas'] = col
            elif 'EXISTENCIA' in col_upper or 'EXISTENCIAS' in col_upper or 'STOCK' in col_upper:
                self.col_mapping['existencias'] = col
            elif 'UBICACIÓN' in col_upper or 'UBICACION' in col_upper:
                self.col_mapping['ubicacion'] = col
            elif 'CADUCIDAD' in col_upper:
                self.col_mapping['caducidad'] = col
            elif 'ANOTACIONES' in col_upper or 'OBSERVACIONES' in col_upper:
                self.col_mapping['anotaciones'] = col
            elif 'TIPO' in col_upper or 'CATEGORIA' in col_upper:
                self.col_mapping['tipo'] = col
            elif 'DONANTE' in col_upper:
                self.col_mapping['donante_col'] = col

        # Fallback por posición
        if len(columnas) >= 1 and 'medicamento' not in self.col_mapping:
            self.col_mapping['medicamento'] = columnas[0]
        if len(columnas) >= 2 and 'presentacion' not in self.col_mapping:
            self.col_mapping['presentacion'] = columnas[1]
        if len(columnas) >= 3 and 'proveedor' not in self.col_mapping:
            self.col_mapping['proveedor'] = columnas[2]
        if len(columnas) >= 4 and 'entradas' not in self.col_mapping:
            self.col_mapping['entradas'] = columnas[3]
        if len(columnas) >= 5 and 'salidas' not in self.col_mapping:
            self.col_mapping['salidas'] = columnas[4]
        if len(columnas) >= 6 and 'existencias' not in self.col_mapping:
            self.col_mapping['existencias'] = columnas[5]
        if len(columnas) >= 7 and 'ubicacion' not in self.col_mapping:
            self.col_mapping['ubicacion'] = columnas[6]
        if len(columnas) >= 8 and 'caducidad' not in self.col_mapping:
            self.col_mapping['caducidad'] = columnas[7]
        if len(columnas) >= 9 and 'anotaciones' not in self.col_mapping:
            self.col_mapping['anotaciones'] = columnas[8]

    def exportar_a_csv_para_verificacion(self, archivo_salida="verificacion_extraccion.csv"):
        """Exporta los datos extraídos a CSV"""
        with open(archivo_salida, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(['=== CATÁLOGO ==='])
            writer.writerow(['nombre', 'forma_farmaceutica', 'concentracion', 'es_controlado'])
            for med in self.medicamentos:
                writer.writerow([med['nombre'], med['forma_farmaceutica'], med['concentracion'], med['es_controlado']])
            writer.writerow([])
            writer.writerow(['=== INVENTARIO ==='])
            writer.writerow(
                ['nombre_medicamento', 'ubicacion', 'caducidad', 'stock', 'es_donacion', 'donante', 'observaciones'])
            for inv in self.inventario:
                writer.writerow(
                    [inv['nombre_medicamento'], inv['ubicacion'], inv['caducidad'], inv['stock'], inv['es_donacion'],
                     inv['donante'], inv['observaciones']])
            writer.writerow([])
            writer.writerow(['=== MOVIMIENTOS ==='])
            writer.writerow(['nombre_medicamento', 'tipo', 'cantidad', 'motivo', 'fecha'])
            for mov in self.movimientos:
                writer.writerow([mov['nombre_medicamento'], mov['tipo'], mov['cantidad'], mov['motivo'], mov['fecha']])
        self._log(f"✅ Datos extraídos guardados en {archivo_salida} para verificación.")

    def extraer_datos(self, nombre_hoja='CONSOLIDADO', generar_log=False, archivo_log="procesamiento_log.txt"):
        """Extrae todos los datos del Excel"""
        if generar_log:
            self.log_file = open(archivo_log, 'w', encoding='utf-8')
            self._log(f"Inicio de procesamiento - {datetime.now()}", "INICIO")
            self._log(f"Archivo Excel: {self.archivo_excel}")
            self._log(f"Hoja: {nombre_hoja}")

        self.omitidos = []
        self._log("\n📂 Cargando archivo Excel...")
        self.wb = load_workbook(self.archivo_excel, data_only=True)

        if nombre_hoja not in self.wb.sheetnames:
            self._log(f"❌ No se encontró la hoja '{nombre_hoja}'")
            self._log(f"   Hojas disponibles: {self.wb.sheetnames}")
            if self.log_file:
                self.log_file.close()
            return False

        sheet_name = nombre_hoja
        self._log(f"📄 Procesando hoja: '{sheet_name}'")
        df_raw = pd.read_excel(self.archivo_excel, sheet_name=sheet_name, header=None)

        # Buscar encabezados
        header_row = None
        for idx in range(min(10, len(df_raw))):
            row = df_raw.iloc[idx]
            if pd.notna(row[0]) and str(row[0]).strip().upper() == 'MEDICAMENTO':
                header_row = idx
                break

        if header_row is None:
            self._log("❌ No se encontró la fila de encabezados (buscando 'MEDICAMENTO')")
            if self.log_file:
                self.log_file.close()
            return False

        self._log(f"📍 Encabezados encontrados en fila Excel: {header_row + 1}")

        encabezados_raw = df_raw.iloc[header_row].tolist()
        encabezados = [str(col).strip() if pd.notna(col) else '' for col in encabezados_raw]

        df = df_raw.iloc[header_row + 1:].reset_index(drop=True)
        df.columns = encabezados

        self._mapear_columnas(df.columns.tolist())

        col_med = self.col_mapping.get('medicamento')
        if col_med is None:
            col_med = df.columns[0]
            self._log(f"⚠️ Usando primera columna como medicamento: '{col_med}'")

        ws = self.wb[sheet_name]
        excel_start_row = header_row + 2

        total_filas = len(df)
        self._log(f"📊 Total de filas en la hoja (después de encabezados): {total_filas}")

        registros_procesados = 0
        registros_omitidos = 0

        self._log("🔍 Procesando cada fila...\n")

        for idx, row in df.iterrows():
            fila_excel_num = excel_start_row + idx
            nombre_val = row[col_med]
            if pd.isna(nombre_val):
                motivo = "Nombre vacío o NaN"
                self.omitidos.append([fila_excel_num, motivo, "", ""])
                registros_omitidos += 1
                continue

            nombre = str(nombre_val).strip()
            if not nombre or nombre.lower() in ('nan', 'none') or len(nombre) < 2:
                motivo = f"Nombre inválido: '{nombre}'"
                self.omitidos.append([fila_excel_num, motivo, nombre, ""])
                registros_omitidos += 1
                continue

            col_pres = self.col_mapping.get('presentacion')
            presentacion_raw = ''
            if col_pres and pd.notna(row[col_pres]):
                presentacion_raw = str(row[col_pres]).strip()
                if presentacion_raw.lower() in ('nan', 'none'):
                    presentacion_raw = ''

            concentracion = self._extraer_concentracion(nombre, presentacion_raw)
            forma_farmaceutica = self._extraer_forma_farmaceutica(presentacion_raw, nombre)
            es_controlado = self._detectar_color_morado(ws, fila_excel_num, columna_letra='A')
            nombre_limpio = self._limpiar_nombre_medicamento(nombre)

            col_ubic = self.col_mapping.get('ubicacion')
            ubicacion = None
            if col_ubic and pd.notna(row[col_ubic]):
                ubicacion = str(row[col_ubic]).strip()
                if ubicacion.lower() in ('nan', 'none'):
                    ubicacion = None

            col_cad = self.col_mapping.get('caducidad')
            caducidad = None
            if col_cad and pd.notna(row[col_cad]):
                caducidad = self._parsear_fecha_caducidad(row[col_cad])

            col_entradas = self.col_mapping.get('entradas')
            entradas = self._extraer_numero(row[col_entradas]) if col_entradas else 0
            col_salidas = self.col_mapping.get('salidas')
            salidas = self._extraer_numero(row[col_salidas]) if col_salidas else 0
            col_stock = self.col_mapping.get('existencias')
            stock_actual = self._extraer_numero(row[col_stock]) if col_stock else 0

            col_anot = self.col_mapping.get('anotaciones')
            observaciones = ''
            if col_anot and pd.notna(row[col_anot]):
                obs = str(row[col_anot]).strip()
                if obs and obs.lower() not in ('nan', 'none'):
                    observaciones = obs

            tipo_registro = ''
            if 'tipo' in self.col_mapping and pd.notna(row[self.col_mapping['tipo']]):
                tipo_registro = str(row[self.col_mapping['tipo']]).strip().upper()

            donante_explicito = None
            if 'donante_col' in self.col_mapping and pd.notna(row[self.col_mapping['donante_col']]):
                donante_explicito = str(row[self.col_mapping['donante_col']]).strip()

            es_donacion = False
            donante = None
            accion_extra = ""

            if tipo_registro == 'SUPLEMENTO':
                if observaciones:
                    observaciones = f"Suplemento. {observaciones}"
                else:
                    observaciones = "Suplemento"
                accion_extra = "→ observaciones actualizadas"
            elif tipo_registro == 'DONACION':
                es_donacion = True
                donante = donante_explicito if donante_explicito else "FARMAMIGO"
                if observaciones:
                    observaciones = f"Donación. {observaciones}"
                else:
                    observaciones = "Donación"
                accion_extra = f"→ donante={donante}"
            else:
                if not es_donacion and observaciones:
                    if any(p in observaciones.lower() for p in ['donacion', 'donación', 'muestra', 'donado']):
                        es_donacion = True
                        donante = donante_explicito or "Donación sin especificar"
                        accion_extra = "→ detectada por palabras clave"

            self.medicamentos.append({
                'nombre': nombre_limpio,
                'forma_farmaceutica': forma_farmaceutica if forma_farmaceutica else None,
                'concentracion': concentracion if concentracion else None,
                'es_controlado': es_controlado
            })

            self.inventario.append({
                'nombre_medicamento': nombre_limpio,
                'ubicacion': ubicacion,
                'caducidad': caducidad,
                'stock': stock_actual,
                'es_donacion': es_donacion,
                'donante': donante,
                'observaciones': observaciones if observaciones else None
            })

            if entradas > 0:
                self.movimientos.append({
                    'nombre_medicamento': nombre_limpio,
                    'tipo': 'entrada',
                    'cantidad': entradas,
                    'motivo': f'Registro inicial desde Excel - Entradas: {entradas} unidades',
                    'fecha': date.today()
                })
            if salidas > 0:
                self.movimientos.append({
                    'nombre_medicamento': nombre_limpio,
                    'tipo': 'salida',
                    'cantidad': salidas,
                    'motivo': f'Registro inicial desde Excel - Salidas: {salidas} unidades',
                    'fecha': date.today()
                })

            stock_calculado = entradas - salidas
            if stock_calculado != stock_actual and stock_actual > 0:
                diferencia = stock_actual - stock_calculado
                if diferencia > 0:
                    self.movimientos.append({
                        'nombre_medicamento': nombre_limpio,
                        'tipo': 'entrada',
                        'cantidad': diferencia,
                        'motivo': f'Ajuste para conciliar inventario (+{diferencia} unidades) - Carga inicial',
                        'fecha': date.today()
                    })
                elif diferencia < 0:
                    self.movimientos.append({
                        'nombre_medicamento': nombre_limpio,
                        'tipo': 'salida',
                        'cantidad': abs(diferencia),
                        'motivo': f'Ajuste para conciliar inventario ({abs(diferencia)} unidades) - Carga inicial',
                        'fecha': date.today()
                    })

            registros_procesados += 1

            if registros_procesados % 50 == 0:
                self._log(f"  Procesados {registros_procesados} registros hasta fila {fila_excel_num}...")

        self._eliminar_duplicados_catalogo()

        mov_entradas = sum(1 for m in self.movimientos if m['tipo'] == 'entrada')
        mov_salidas = sum(1 for m in self.movimientos if m['tipo'] == 'salida')

        self._log(f"\n✅ Procesamiento completado:")
        self._log(f"  - Total filas leídas en hoja: {total_filas}")
        self._log(f"  - Registros procesados (incluidos): {registros_procesados}")
        self._log(f"  - Registros omitidos: {registros_omitidos}")
        self._log(f"  - Medicamentos únicos en catálogo: {len(self.medicamentos)}")
        self._log(f"  - Registros de inventario: {len(self.inventario)}")
        self._log(
            f"  - Movimientos generados: {len(self.movimientos)} (entradas: {mov_entradas}, salidas: {mov_salidas})")

        if self.omitidos:
            with open("omitidos.csv", 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow(["Fila Excel", "Motivo", "Nombre", "Presentación"])
                writer.writerows(self.omitidos)
            self._log(f"\n⚠️ Se omitieron {len(self.omitidos)} filas. Revisa el archivo 'omitidos.csv' para detalles.")
        else:
            self._log(f"\n✅ No se omitió ninguna fila.")

        if self.log_file:
            self._log(f"\n📄 Log detallado guardado en {archivo_log}")
            self.log_file.close()
            self.log_file = None

        return True

    def reiniciar_contadores_identidad(self, cursor):
        """Reinicia los contadores de identidad"""
        try:
            self._log("  🔄 Reiniciando contadores de identidad...")
            cursor.execute("DBCC CHECKIDENT ('catalogo_medicamentos', RESEED, 0)")
            self._log("     ✅ ID de catálogo reiniciado a 1")
            cursor.execute("DBCC CHECKIDENT ('inventario_farmacia', RESEED, 0)")
            self._log("     ✅ ID de inventario reiniciado a 1")
            cursor.execute("DBCC CHECKIDENT ('movimientos_farmacia', RESEED, 0)")
            self._log("     ✅ ID de movimientos reiniciado a 1")
        except pyodbc.Error as e:
            self._log(f"     ⚠️ Error al reiniciar contadores: {e}")

    def insertar_en_base_datos(self, server, database, limpiar_tablas=False, reiniciar_ids=True, username=None,
                               password=None):
        """Inserta los datos directamente en la base de datos"""
        self._log("\n💾 Conectando a la base de datos...")

        if username and password:
            conn_str = f'DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={server};DATABASE={database};UID={username};PWD={password};'
        else:
            conn_str = f'DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={server};DATABASE={database};Trusted_Connection=yes;'

        try:
            conn = pyodbc.connect(conn_str, timeout=30)
            cursor = conn.cursor()
            self._log("✅ Conexión exitosa!")

            self._obtener_usuario_admin(cursor)

            if limpiar_tablas:
                self._log("\n🗑️ Limpiando tablas existentes...")
                cursor.execute("DELETE FROM movimientos_farmacia")
                self._log("  ✅ Movimientos eliminados")
                cursor.execute("DELETE FROM inventario_farmacia")
                self._log("  ✅ Inventario eliminado")
                cursor.execute("DELETE FROM catalogo_medicamentos")
                self._log("  ✅ Catálogo eliminado")

                if reiniciar_ids:
                    self.reiniciar_contadores_identidad(cursor)

            self._log("\n📝 Insertando catálogo de medicamentos...")
            catalogos_insertados = 0
            for med in self.medicamentos:
                cursor.execute("""
                    INSERT INTO catalogo_medicamentos (nombre, forma_farmaceutica, concentracion, es_controlado)
                    VALUES (?, ?, ?, ?)
                """, (med['nombre'], med['forma_farmaceutica'], med['concentracion'], med['es_controlado']))
                catalogos_insertados += 1
                if catalogos_insertados % 50 == 0:
                    self._log(f"  Insertados {catalogos_insertados} medicamentos...")
            self._log(f"  ✅ {catalogos_insertados} medicamentos insertados")

            self._log("\n📦 Insertando inventario...")
            inventario_insertados = 0
            for inv in self.inventario:
                cursor.execute("SELECT id_catalogo FROM catalogo_medicamentos WHERE nombre = ?",
                               (inv['nombre_medicamento'],))
                row = cursor.fetchone()
                if row:
                    id_catalogo = row[0]
                    cursor.execute("""
                        INSERT INTO inventario_farmacia (id_catalogo, ubicacion, caducidad, stock, es_donacion, donante, observaciones)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    """, (id_catalogo, inv['ubicacion'], inv['caducidad'], inv['stock'],
                          inv['es_donacion'], inv['donante'], inv['observaciones']))
                    inventario_insertados += 1
                    if inventario_insertados % 50 == 0:
                        self._log(f"  Insertados {inventario_insertados} registros de inventario...")
            self._log(f"  ✅ {inventario_insertados} registros de inventario insertados")

            self._log("\n📋 Insertando movimientos...")
            movimientos_insertados = 0
            mov_entradas = 0
            mov_salidas = 0

            for mov in self.movimientos:
                cursor.execute("""
                    SELECT TOP 1 i.id_inventario
                    FROM inventario_farmacia i
                    INNER JOIN catalogo_medicamentos c ON i.id_catalogo = c.id_catalogo
                    WHERE c.nombre = ?
                """, (mov['nombre_medicamento'],))
                row = cursor.fetchone()
                if row:
                    id_inventario = row[0]
                    if mov['tipo'] == 'entrada':
                        quien_recibe = self.usuario_admin_nombre
                        quien_retira = None
                    else:
                        quien_recibe = None
                        quien_retira = self.usuario_admin_nombre

                    cursor.execute("""
                        INSERT INTO movimientos_farmacia (id_inventario, tipo, cantidad, quien_recibe, quien_retira, motivo, fecha)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    """, (id_inventario, mov['tipo'], mov['cantidad'], quien_recibe, quien_retira, mov['motivo'],
                          mov['fecha']))
                    movimientos_insertados += 1
                    if mov['tipo'] == 'entrada':
                        mov_entradas += 1
                    else:
                        mov_salidas += 1

                    if movimientos_insertados % 100 == 0:
                        self._log(f"  Insertados {movimientos_insertados} movimientos...")

            self._log(f"  ✅ {movimientos_insertados} movimientos insertados")
            self._log(f"     · Entradas: {mov_entradas}")
            self._log(f"     · Salidas: {mov_salidas}")
            self._log(f"     · Asignados al usuario: {self.usuario_admin_nombre}")

            conn.commit()
            cursor.close()
            conn.close()

            self._log("\n" + "=" * 70)
            self._log("✅ ¡ÉXITO! Todos los datos fueron insertados correctamente")
            self._log("=" * 70)
            return True

        except pyodbc.Error as e:
            self._log(f"\n❌ Error de base de datos: {e}")
            return False
        except Exception as e:
            self._log(f"\n❌ Error inesperado: {e}")
            import traceback
            traceback.print_exc()
            return False


# ==================== INTERFAZ GRÁFICA ====================
class AplicacionFarmacia:
    def __init__(self, root):
        self.root = root
        self.root.title("Importador de Datos - Farmacia v1.0")
        self.root.geometry("950x750")
        self.root.resizable(True, True)

        # Variable para el extractor
        self.extractor = None

        # Configurar estilos
        self.setup_styles()

        # Crear interfaz
        self.create_widgets()

        # Configurar manejo de cierre
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

    def setup_styles(self):
        """Configurar colores y estilos"""
        self.colors = {
            'bg': '#f5f5f5',
            'header': '#2c3e50',
            'button': '#3498db',
            'button_hover': '#2980b9',
            'success': '#27ae60',
            'warning': '#e74c3c',
            'info': '#34495e',
            'frame_bg': '#ffffff'
        }

        self.root.configure(bg=self.colors['bg'])

        # Fuentes
        self.font_title = tkfont.Font(family="Segoe UI", size=14, weight="bold")
        self.font_normal = tkfont.Font(family="Segoe UI", size=10)
        self.font_bold = tkfont.Font(family="Segoe UI", size=10, weight="bold")

    def create_widgets(self):
        # Frame principal con scroll
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Canvas y scrollbar para scroll
        canvas = tk.Canvas(main_frame, bg=self.colors['bg'], highlightthickness=0)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg=self.colors['bg'])

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        title_frame = tk.Frame(scrollable_frame, bg=self.colors['header'], height=60)
        title_frame.pack(fill=tk.X, pady=(0, 10))
        title_frame.pack_propagate(False)

        title_label = tk.Label(title_frame, text="🏥 IMPORTADOR DE DATOS - SISTEMA DE FARMACIA",
                               font=self.font_title, bg=self.colors['header'], fg='white')
        title_label.pack(expand=True)

        # ========== SECCIÓN ARCHIVO ==========
        file_frame = tk.LabelFrame(scrollable_frame, text="📁 Archivo Excel",
                                   font=self.font_bold, bg=self.colors['frame_bg'], fg=self.colors['header'])
        file_frame.pack(fill=tk.X, padx=10, pady=10)

        tk.Label(file_frame, text="Archivo:", font=self.font_normal, bg=self.colors['frame_bg']).grid(row=0, column=0,
                                                                                                      padx=5, pady=10,
                                                                                                      sticky='e')
        self.file_path = tk.StringVar()
        tk.Entry(file_frame, textvariable=self.file_path, width=60, state='readonly',
                 font=self.font_normal).grid(row=0, column=1, padx=5, pady=10)
        tk.Button(file_frame, text="📂 Seleccionar", command=self.seleccionar_archivo,
                  bg=self.colors['button'], fg='white', cursor='hand2',
                  font=self.font_normal, padx=10).grid(row=0, column=2, padx=5, pady=10)

        # ========== SECCIÓN CONFIGURACIÓN ==========
        config_frame = tk.LabelFrame(scrollable_frame, text="⚙️ Configuración de Extracción",
                                     font=self.font_bold, bg=self.colors['frame_bg'], fg=self.colors['header'])
        config_frame.pack(fill=tk.X, padx=10, pady=10)

        tk.Label(config_frame, text="Nombre de hoja:", font=self.font_normal, bg=self.colors['frame_bg']).grid(row=0,
                                                                                                               column=0,
                                                                                                               padx=5,
                                                                                                               pady=5,
                                                                                                               sticky='e')
        self.sheet_name = tk.StringVar(value="CONSOLIDADO")
        tk.Entry(config_frame, textvariable=self.sheet_name, width=30, font=self.font_normal).grid(row=0, column=1,
                                                                                                   padx=5, pady=5,
                                                                                                   sticky='w')

        self.generar_log = tk.BooleanVar(value=True)
        tk.Checkbutton(config_frame, text="Generar log detallado del procesamiento", variable=self.generar_log,
                       bg=self.colors['frame_bg'], font=self.font_normal).grid(row=0, column=2, padx=5, pady=5,
                                                                               sticky='w')

        # ========== SECCIÓN BASE DE DATOS ==========
        db_frame = tk.LabelFrame(scrollable_frame, text="💾 Configuración de Base de Datos",
                                 font=self.font_bold, bg=self.colors['frame_bg'], fg=self.colors['header'])
        db_frame.pack(fill=tk.X, padx=10, pady=10)

        # Servidor
        tk.Label(db_frame, text="Servidor:", font=self.font_normal, bg=self.colors['frame_bg']).grid(row=0, column=0,
                                                                                                     padx=5, pady=5,
                                                                                                     sticky='e')
        self.server = tk.StringVar()
        tk.Entry(db_frame, textvariable=self.server, width=35, font=self.font_normal).grid(row=0, column=1, padx=5,
                                                                                           pady=5)

        # Base de datos
        tk.Label(db_frame, text="Base de datos:", font=self.font_normal, bg=self.colors['frame_bg']).grid(row=0,
                                                                                                          column=2,
                                                                                                          padx=5,
                                                                                                          pady=5,
                                                                                                          sticky='e')
        self.database = tk.StringVar()
        tk.Entry(db_frame, textvariable=self.database, width=35, font=self.font_normal).grid(row=0, column=3, padx=5,
                                                                                             pady=5)

        # Separador
        ttk.Separator(db_frame, orient='horizontal').grid(row=1, column=0, columnspan=4, sticky='ew', pady=10)

        # Autenticación
        self.auth_type = tk.StringVar(value="windows")
        tk.Radiobutton(db_frame, text="Autenticación de Windows", variable=self.auth_type,
                       value="windows", bg=self.colors['frame_bg'], font=self.font_normal).grid(row=2, column=0,
                                                                                                columnspan=2, padx=5,
                                                                                                pady=5, sticky='w')
        tk.Radiobutton(db_frame, text="Autenticación SQL Server", variable=self.auth_type,
                       value="sql", bg=self.colors['frame_bg'], font=self.font_normal).grid(row=2, column=2,
                                                                                            columnspan=2, padx=5,
                                                                                            pady=5, sticky='w')

        # Usuario y contraseña (inicialmente deshabilitados)
        self.username = tk.StringVar()
        self.password = tk.StringVar()
        self.username_entry = tk.Entry(db_frame, textvariable=self.username, width=30, font=self.font_normal,
                                       state='disabled')
        self.password_entry = tk.Entry(db_frame, textvariable=self.password, show="*", width=30, font=self.font_normal,
                                       state='disabled')

        tk.Label(db_frame, text="Usuario:", font=self.font_normal, bg=self.colors['frame_bg']).grid(row=3, column=0,
                                                                                                    padx=5, pady=5,
                                                                                                    sticky='e')
        self.username_entry.grid(row=3, column=1, padx=5, pady=5)
        tk.Label(db_frame, text="Contraseña:", font=self.font_normal, bg=self.colors['frame_bg']).grid(row=3, column=2,
                                                                                                       padx=5, pady=5,
                                                                                                       sticky='e')
        self.password_entry.grid(row=3, column=3, padx=5, pady=5)

        # Evento para habilitar/deshabilitar campos según autenticación
        self.auth_type.trace('w', self.toggle_auth_fields)

        # Separador
        ttk.Separator(db_frame, orient='horizontal').grid(row=4, column=0, columnspan=4, sticky='ew', pady=10)

        # Opciones
        self.limpiar_tablas = tk.BooleanVar(value=False)
        self.reiniciar_ids = tk.BooleanVar(value=False)

        tk.Checkbutton(db_frame, text="Limpiar tablas existentes antes de insertar", variable=self.limpiar_tablas,
                       bg=self.colors['frame_bg'], font=self.font_normal, command=self.toggle_reiniciar_ids).grid(row=5,
                                                                                                                  column=0,
                                                                                                                  columnspan=2,
                                                                                                                  padx=5,
                                                                                                                  pady=5,
                                                                                                                  sticky='w')

        self.reiniciar_check = tk.Checkbutton(db_frame, text="Reiniciar IDs (empezar desde 1)",
                                              variable=self.reiniciar_ids, bg=self.colors['frame_bg'],
                                              font=self.font_normal, state='disabled')
        self.reiniciar_check.grid(row=5, column=2, columnspan=2, padx=5, pady=5, sticky='w')

        # ========== BOTONES DE ACCIÓN ==========
        button_frame = tk.Frame(scrollable_frame, bg=self.colors['bg'])
        button_frame.pack(fill=tk.X, padx=10, pady=10)

        self.btn_extraer = tk.Button(button_frame, text="📊 1. Extraer Datos", command=self.extraer_datos,
                                     bg=self.colors['button'], fg='white', font=self.font_bold,
                                     cursor='hand2', height=2, width=18)
        self.btn_extraer.pack(side=tk.LEFT, padx=5)

        self.btn_verificar = tk.Button(button_frame, text="📄 2. Verificar CSV", command=self.verificar_datos,
                                       bg=self.colors['info'], fg='white', font=self.font_bold,
                                       cursor='hand2', height=2, width=18, state='disabled')
        self.btn_verificar.pack(side=tk.LEFT, padx=5)

        self.btn_insertar = tk.Button(button_frame, text="💾 3. Insertar en BD", command=self.insertar_datos,
                                      bg=self.colors['success'], fg='white', font=self.font_bold,
                                      cursor='hand2', height=2, width=18, state='disabled')
        self.btn_insertar.pack(side=tk.LEFT, padx=5)

        tk.Button(button_frame, text="🗑️ Limpiar Log", command=self.limpiar_log,
                  bg=self.colors['warning'], fg='white', font=self.font_bold,
                  cursor='hand2', height=2, width=15).pack(side=tk.RIGHT, padx=5)

        # ========== LOG Y PROGRESO ==========
        log_frame = tk.LabelFrame(scrollable_frame, text="📋 Registro de Actividades",
                                  font=self.font_bold, bg=self.colors['frame_bg'])
        log_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        self.log_text = scrolledtext.ScrolledText(log_frame, height=15, width=80,
                                                  font=("Consolas", 9), wrap=tk.WORD)
        self.log_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Barra de progreso
        self.progress = ttk.Progressbar(scrollable_frame, mode='indeterminate')
        self.progress.pack(fill=tk.X, padx=10, pady=5)

        # Barra de estado
        self.status_bar = tk.Label(scrollable_frame, text="Listo", bd=1, relief=tk.SUNKEN, anchor=tk.W,
                                   font=self.font_normal, bg=self.colors['info'], fg='white')
        self.status_bar.pack(fill=tk.X, padx=10, pady=(0, 5))

        # Empaquetar canvas y scrollbar
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

    def toggle_auth_fields(self, *args):
        """Habilitar/deshabilitar campos de autenticación"""
        if self.auth_type.get() == "sql":
            self.username_entry.config(state='normal')
            self.password_entry.config(state='normal')
        else:
            self.username_entry.config(state='disabled')
            self.password_entry.config(state='disabled')
            self.username.set("")
            self.password.set("")

    def toggle_reiniciar_ids(self):
        """Habilitar/deshabilitar opción de reiniciar IDs"""
        if self.limpiar_tablas.get():
            self.reiniciar_check.config(state='normal')
        else:
            self.reiniciar_check.config(state='disabled')
            self.reiniciar_ids.set(False)

    def seleccionar_archivo(self):
        """Seleccionar archivo Excel"""
        archivo = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Archivos Excel", "*.xlsx *.xls"), ("Todos los archivos", "*.*")]
        )
        if archivo:
            self.file_path.set(archivo)
            self.log(f"✅ Archivo seleccionado: {os.path.basename(archivo)}")
            self.update_status(f"Archivo cargado: {os.path.basename(archivo)}")

    def log(self, mensaje, nivel="INFO"):
        """Agregar mensaje al log"""
        timestamp = datetime.now().strftime("%H:%M:%S")

        # Colorear según nivel
        tag = None
        if "✅" in mensaje or "ÉXITO" in mensaje:
            tag = "success"
        elif "❌" in mensaje or "Error" in mensaje:
            tag = "error"
        elif "⚠️" in mensaje:
            tag = "warning"
        elif "📊" in mensaje or "📦" in mensaje or "📋" in mensaje:
            tag = "info"

        self.log_text.insert(tk.END, f"[{timestamp}] {mensaje}\n", tag)
        self.log_text.see(tk.END)

        # Configurar tags de color
        self.log_text.tag_config("success", foreground="green")
        self.log_text.tag_config("error", foreground="red")
        self.log_text.tag_config("warning", foreground="orange")
        self.log_text.tag_config("info", foreground="blue")

        self.root.update_idletasks()

    def update_status(self, mensaje):
        """Actualizar barra de estado"""
        self.status_bar.config(text=mensaje)
        self.root.update_idletasks()

    def limpiar_log(self):
        """Limpiar el área de log"""
        self.log_text.delete(1.0, tk.END)
        self.log("🧹 Log limpiado")

    def extraer_datos(self):
        """Extraer datos del Excel en un hilo separado"""
        if not self.file_path.get():
            messagebox.showerror("Error", "Por favor seleccione un archivo Excel")
            return

        # Deshabilitar botones durante la extracción
        self.btn_extraer.config(state='disabled')
        self.btn_verificar.config(state='disabled')
        self.btn_insertar.config(state='disabled')
        self.progress.start()
        self.update_status("Extrayendo datos del Excel...")

        # Ejecutar en hilo separado
        thread = threading.Thread(target=self._extraer_datos_thread)
        thread.daemon = True
        thread.start()

    def _extraer_datos_thread(self):
        """Thread para extraer datos"""
        try:
            self.log("🔍 Iniciando extracción de datos...")

            # Crear extractor
            self.extractor = ExtraerDatosFarmacia(self.file_path.get())

            # Conectar callback de log
            self.extractor.log_callback = self.log

            # Extraer datos
            success = self.extractor.extraer_datos(
                nombre_hoja=self.sheet_name.get(),
                generar_log=self.generar_log.get(),
                archivo_log="procesamiento_log.txt"
            )

            if success:
                self.log(f"✅ Extracción completada exitosamente")
                self.log(f"   📊 Medicamentos únicos: {len(self.extractor.medicamentos)}")
                self.log(f"   📦 Registros inventario: {len(self.extractor.inventario)}")
                self.log(f"   📋 Movimientos generados: {len(self.extractor.movimientos)}")

                # Habilitar botones siguientes
                self.btn_verificar.config(state='normal')
                self.btn_insertar.config(state='normal')

                self.update_status(f"Extracción completada - {len(self.extractor.medicamentos)} medicamentos")
                messagebox.showinfo("Éxito",
                                    f"Datos extraídos correctamente\n\nMedicamentos: {len(self.extractor.medicamentos)}\nInventario: {len(self.extractor.inventario)}\nMovimientos: {len(self.extractor.movimientos)}")
            else:
                self.log("❌ Error en la extracción de datos")
                self.update_status("Error en la extracción")
                messagebox.showerror("Error", "No se pudieron extraer los datos. Revise el log.")

        except Exception as e:
            self.log(f"❌ Error: {str(e)}")
            self.update_status(f"Error: {str(e)[:50]}")
            messagebox.showerror("Error", f"Error durante la extracción:\n{str(e)}")
        finally:
            self.progress.stop()
            self.btn_extraer.config(state='normal')

    def verificar_datos(self):
        """Exportar datos a CSV para verificación"""
        if not self.extractor:
            messagebox.showerror("Error", "Primero debe extraer los datos")
            return

        archivo = filedialog.asksaveasfilename(
            title="Guardar CSV de verificación",
            defaultextension=".csv",
            filetypes=[("Archivos CSV", "*.csv")],
            initialfile="verificacion_datos.csv"
        )

        if archivo:
            try:
                self.update_status("Exportando a CSV...")
                self.extractor.exportar_a_csv_para_verificacion(archivo)
                self.log(f"✅ Datos exportados a: {archivo}")
                self.update_status(f"CSV guardado: {os.path.basename(archivo)}")
                messagebox.showinfo("Éxito", f"Datos exportados correctamente a:\n{archivo}")
            except Exception as e:
                self.log(f"❌ Error al exportar: {str(e)}")
                self.update_status("Error al exportar CSV")
                messagebox.showerror("Error", f"Error al exportar:\n{str(e)}")

    def insertar_datos(self):
        """Insertar datos en la base de datos"""
        if not self.extractor:
            messagebox.showerror("Error", "Primero debe extraer los datos")
            return

        # Validar conexión
        if not self.server.get() or not self.database.get():
            messagebox.showerror("Error", "Debe especificar servidor y base de datos")
            return

        # Confirmar inserción
        mov_count = len(self.extractor.movimientos)
        if not messagebox.askyesno("Confirmar Inserción",
                                   f"¿Desea insertar los datos en la base de datos?\n\n"
                                   f"📊 Resumen:\n"
                                   f"  • Medicamentos: {len(self.extractor.medicamentos)}\n"
                                   f"  • Inventario: {len(self.extractor.inventario)}\n"
                                   f"  • Movimientos: {mov_count}\n\n"
                                   f"Servidor: {self.server.get()}\n"
                                   f"Base de datos: {self.database.get()}\n\n"
                                   f"¿Continuar?"):
            return

        # Deshabilitar botones
        self.btn_insertar.config(state='disabled')
        self.btn_verificar.config(state='disabled')
        self.progress.start()
        self.update_status("Insertando datos en la base de datos...")

        # Ejecutar en hilo
        thread = threading.Thread(target=self._insertar_datos_thread)
        thread.daemon = True
        thread.start()

    def _insertar_datos_thread(self):
        """Thread para insertar datos"""
        try:
            self.log("💾 Conectando a la base de datos...")

            # Preparar parámetros
            username = self.username.get() if self.auth_type.get() == "sql" else None
            password = self.password.get() if self.auth_type.get() == "sql" else None

            # Insertar
            success = self.extractor.insertar_en_base_datos(
                server=self.server.get(),
                database=self.database.get(),
                limpiar_tablas=self.limpiar_tablas.get(),
                reiniciar_ids=self.reiniciar_ids.get(),
                username=username,
                password=password
            )

            if success:
                self.log("✅ ¡Datos insertados exitosamente en la base de datos!")
                self.update_status("Inserción completada con éxito")
                messagebox.showinfo("Éxito", "Datos insertados correctamente en la base de datos")
            else:
                self.log("❌ Error al insertar datos")
                self.update_status("Error en la inserción")
                messagebox.showerror("Error", "Error al insertar datos. Revise el log para más detalles")

        except Exception as e:
            self.log(f"❌ Error: {str(e)}")
            self.update_status(f"Error: {str(e)[:50]}")
            messagebox.showerror("Error", f"Error durante la inserción:\n{str(e)}")
        finally:
            self.progress.stop()
            self.btn_insertar.config(state='normal')
            self.btn_verificar.config(state='normal')

    def on_closing(self):
        """Manejar cierre de la aplicación"""
        if messagebox.askokcancel("Salir", "¿Desea salir de la aplicación?"):
            self.root.destroy()


# ==================== MAIN ====================
def main():
    root = tk.Tk()
    app = AplicacionFarmacia(root)
    root.mainloop()


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"Error: {e}")
        import traceback

        traceback.print_exc()
        input("Presione Enter para salir...")