import pandas as pd
import openpyxl
from openpyxl import load_workbook
import re
from datetime import datetime, date
from calendar import monthrange
import os
import pyodbc
import csv


class ExtraerDatosFarmacia:
    def __init__(self, archivo_excel):
        self.archivo_excel = archivo_excel
        self.medicamentos = []
        self.inventario = []
        self.movimientos = []
        self.wb = None
        self.col_mapping = {}
        self.usuario_admin_nombre = None
        self.usuario_admin_id = None
        self.log_file = None          # Archivo de log (se abre si se pide)
        self.omitidos = []            # Lista para guardar filas omitidas

    # ==================== MÉTODOS AUXILIARES ====================
    def _log(self, mensaje, nivel="INFO"):
        """Escribe en el archivo de log si está abierto, y también en consola"""
        if self.log_file:
            self.log_file.write(f"[{nivel}] {mensaje}\n")
            self.log_file.flush()
        print(mensaje)

    def _obtener_usuario_admin(self, cursor):
        """Obtiene el usuario administrador (nombre_usuario = 'admin') de la base de datos"""
        try:
            cursor.execute("""
                SELECT id_usuario, nombre_usuario
                FROM usuarios
                WHERE nombre_usuario = 'admin'
            """)
            row = cursor.fetchone()
            if row:
                self.usuario_admin_id = row[0]
                self.usuario_admin_nombre = row[1]
                print(f"  👤 Usuario admin encontrado: {self.usuario_admin_nombre} (ID: {self.usuario_admin_id})")
            else:
                cursor.execute("SELECT TOP 1 id_usuario, nombre_usuario FROM usuarios ORDER BY id_usuario")
                row = cursor.fetchone()
                if row:
                    self.usuario_admin_id = row[0]
                    self.usuario_admin_nombre = row[1]
                    print(f"  👤 Usuario por defecto: {self.usuario_admin_nombre} (ID: {self.usuario_admin_id})")
                else:
                    self.usuario_admin_nombre = "admin"
                    self.usuario_admin_id = None
                    print(f"  ⚠️ No se encontraron usuarios, se usará 'admin'")
            return True
        except Exception as e:
            print(f"  ⚠️ Error al obtener usuario admin: {e}")
            self.usuario_admin_nombre = "admin"
            self.usuario_admin_id = None
            return True

    def _parsear_fecha_caducidad(self, valor):
        """Parsea fecha de caducidad (puede ser solo mes/año o fecha completa)"""
        if pd.isna(valor):
            return None
        try:
            if isinstance(valor, datetime):
                return valor.date()
            elif isinstance(valor, str):
                valor = valor.strip()
                if valor.lower() in ('nan', 'none', ''):
                    return None
                for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y']:
                    try:
                        return datetime.strptime(valor, fmt).date()
                    except:
                        continue
                # Solo mes/año (ej: "12/2024" o "12-2024")
                patron_mes_anio = r'(\d{1,2})[/-](\d{4})'
                match = re.search(patron_mes_anio, valor)
                if match:
                    mes = int(match.group(1))
                    anio = int(match.group(2))
                    if mes > 12:
                        mes, anio = anio, mes
                    ultimo_dia = monthrange(anio, mes)[1]
                    return date(anio, mes, ultimo_dia)
                # Solo año
                patron_anio = r'(\d{4})'
                match = re.search(patron_anio, valor)
                if match:
                    anio = int(match.group(1))
                    return date(anio, 12, 31)
            elif hasattr(valor, 'date'):
                return valor.date()
        except Exception as e:
            print(f"  ⚠️ Error al parsear fecha: {valor} - {e}")
        return None

    def _extraer_numero(self, valor):
        """Extrae un número entero de un valor que puede ser int, float o string"""
        if pd.isna(valor):
            return 0
        try:
            if isinstance(valor, (int, float)):
                return int(valor)
            elif isinstance(valor, str):
                numeros = re.findall(r'\d+', valor)
                if numeros:
                    return sum(int(n) for n in numeros)
        except:
            pass
        return 0

    def _detectar_color_morado(self, ws, fila, columna_letra='A'):
        """Detecta color morado de fondo en la celda"""
        try:
            celda = ws[f"{columna_letra}{fila}"]
            if celda.fill and hasattr(celda.fill, 'fgColor') and celda.fill.fgColor:
                fg_color = celda.fill.fgColor
                if hasattr(fg_color, 'rgb') and fg_color.rgb:
                    rgb = str(fg_color.rgb).upper()
                    if len(rgb) >= 6:
                        if len(rgb) == 8:
                            rgb = rgb[2:]
                        try:
                            r = int(rgb[0:2], 16)
                            g = int(rgb[2:4], 16)
                            b = int(rgb[4:6], 16)
                            if (r > 100 and b > 100 and g < 150) or (r > 150 and b > 150):
                                return True
                        except:
                            pass
        except:
            pass
        return False

    def _extraer_concentracion(self, nombre, presentacion):
        """Extrae la concentración del medicamento"""
        texto = f"{nombre} {presentacion}".lower()
        patrones = [
            r'(\d+(?:\.\d+)?\s*(?:mg|mcg|µg|g|ml|UI)(?:\s*\/\s*(?:ml|mg|g))?)',
            r'(\d+(?:\.\d+)?\s*mg(?:\s*\/\s*\d+(?:\.\d+)?\s*ml)?)',
            r'(\d+(?:\.\d+)?\s*(?:%|por ciento))',
        ]
        for patron in patrones:
            coincidencia = re.search(patron, texto, re.IGNORECASE)
            if coincidencia:
                return coincidencia.group(1)
        return None

    def _extraer_forma_farmaceutica(self, presentacion, nombre):
        """Extrae la forma farmacéutica del medicamento"""
        texto = f"{presentacion} {nombre}".lower()
        formas = {
            'Tableta': ['tableta', 'tabletas', 'comprimido', 'comprimidos', 'tab', 'tabs'],
            'Cápsula': ['cápsula', 'capsula', 'cápsulas', 'capsulas', 'cap'],
            'Jarabe': ['jarabe', 'solución oral', 'suspensión oral', 'suspension oral'],
            'Inyectable': ['inyectable', 'ampolleta', 'ampolla', 'vial', 'jeringa'],
            'Crema': ['crema', 'ungüento', 'pomada', 'gel'],
            'Polvo': ['polvo', 'granulado', 'sobres'],
            'Gotas': ['gotas', 'gotero'],
            'Aerosol': ['aerosol', 'inhalador', 'spray'],
            'Solución': ['solución', 'solucion', 'solución inyectable'],
        }
        for forma, palabras in formas.items():
            for palabra in palabras:
                if palabra in texto:
                    return forma
        return None

    def _limpiar_nombre_medicamento(self, nombre):
        """Limpia el nombre eliminando concentración y forma farmacéutica"""
        nombre_limpio = re.sub(r'\s+\d+(?:\.\d+)?\s*(?:mg|mcg|µg|g|ml|UI).*$', '', nombre, flags=re.IGNORECASE)
        nombre_limpio = re.sub(r'\s*\(.*?\)\s*$', '', nombre_limpio)
        palabras_eliminar = ['tableta', 'tabletas', 'cápsula', 'capsula', 'comprimido', 'comprimidos',
                             'gragea', 'grageas', 'sobre', 'sobres', 'vial', 'ampolla']
        for palabra in palabras_eliminar:
            if nombre_limpio.lower().endswith(palabra):
                nombre_limpio = nombre_limpio[:-len(palabra)].strip()
                break
        if nombre_limpio and len(nombre_limpio) > 1:
            nombre_limpio = nombre_limpio[0].upper() + nombre_limpio[1:].lower()
        elif nombre_limpio:
            nombre_limpio = nombre_limpio.upper()
        return nombre_limpio.strip()

    def _eliminar_duplicados_catalogo(self):
        """Elimina medicamentos duplicados del catálogo"""
        medicamentos_unicos = []
        claves_vistas = set()
        for med in self.medicamentos:
            clave = f"{med['nombre'].lower()}|{med.get('forma_farmaceutica', '')}|{med.get('concentracion', '')}"
            if clave not in claves_vistas:
                claves_vistas.add(clave)
                medicamentos_unicos.append(med)
        self.medicamentos = medicamentos_unicos

    def _mapear_columnas(self, columnas):
        """Mapea los nombres de columnas (por nombre o por posición esperada)"""
        self.col_mapping = {}
        for i, col in enumerate(columnas):
            col_upper = col.upper() if col else ''
            if 'MEDICAMENTO' in col_upper:
                self.col_mapping['medicamento'] = col
            elif 'PRESENTACIÓN' in col_upper or 'PRESENTACION' in col_upper:
                self.col_mapping['presentacion'] = col
            elif 'PROVEEDOR' in col_upper:
                self.col_mapping['proveedor'] = col
            elif 'ENTRADA' in col_upper or 'ENTRADAS' in col_upper:
                self.col_mapping['entradas'] = col
            elif 'SALIDA' in col_upper or 'SALIDAS' in col_upper:
                self.col_mapping['salidas'] = col
            elif 'EXISTENCIA' in col_upper or 'EXISTENCIAS' in col_upper or 'STOCK' in col_upper:
                self.col_mapping['existencias'] = col
            elif 'UBICACIÓN' in col_upper or 'UBICACION' in col_upper:
                self.col_mapping['ubicacion'] = col
            elif 'CADUCIDAD' in col_upper:
                self.col_mapping['caducidad'] = col
            elif 'ANOTACIONES' in col_upper or 'OBSERVACIONES' in col_upper:
                self.col_mapping['anotaciones'] = col
            elif 'TIPO' in col_upper or 'CATEGORIA' in col_upper:
                self.col_mapping['tipo'] = col
            elif 'DONANTE' in col_upper:
                self.col_mapping['donante_col'] = col

        # Fallback por posición si no se encontraron
        if len(columnas) >= 1 and 'medicamento' not in self.col_mapping:
            self.col_mapping['medicamento'] = columnas[0]
        if len(columnas) >= 2 and 'presentacion' not in self.col_mapping:
            self.col_mapping['presentacion'] = columnas[1]
        if len(columnas) >= 3 and 'proveedor' not in self.col_mapping:
            self.col_mapping['proveedor'] = columnas[2]
        if len(columnas) >= 4 and 'entradas' not in self.col_mapping:
            self.col_mapping['entradas'] = columnas[3]
        if len(columnas) >= 5 and 'salidas' not in self.col_mapping:
            self.col_mapping['salidas'] = columnas[4]
        if len(columnas) >= 6 and 'existencias' not in self.col_mapping:
            self.col_mapping['existencias'] = columnas[5]
        if len(columnas) >= 7 and 'ubicacion' not in self.col_mapping:
            self.col_mapping['ubicacion'] = columnas[6]
        if len(columnas) >= 8 and 'caducidad' not in self.col_mapping:
            self.col_mapping['caducidad'] = columnas[7]
        if len(columnas) >= 9 and 'anotaciones' not in self.col_mapping:
            self.col_mapping['anotaciones'] = columnas[8]

    # ==================== EXPORTACIÓN A CSV PARA VERIFICACIÓN ====================
    def exportar_a_csv_para_verificacion(self, archivo_salida="verificacion_extraccion.csv"):
        """Exporta los datos extraídos (catálogo, inventario, movimientos) a CSV para comparar"""
        with open(archivo_salida, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(['=== CATÁLOGO ==='])
            writer.writerow(['nombre', 'forma_farmaceutica', 'concentracion', 'es_controlado'])
            for med in self.medicamentos:
                writer.writerow([med['nombre'], med['forma_farmaceutica'], med['concentracion'], med['es_controlado']])
            writer.writerow([])
            writer.writerow(['=== INVENTARIO ==='])
            writer.writerow(['nombre_medicamento', 'ubicacion', 'caducidad', 'stock', 'es_donacion', 'donante', 'observaciones'])
            for inv in self.inventario:
                writer.writerow([inv['nombre_medicamento'], inv['ubicacion'], inv['caducidad'], inv['stock'], inv['es_donacion'], inv['donante'], inv['observaciones']])
            writer.writerow([])
            writer.writerow(['=== MOVIMIENTOS ==='])
            writer.writerow(['nombre_medicamento', 'tipo', 'cantidad', 'motivo', 'fecha'])
            for mov in self.movimientos:
                writer.writerow([mov['nombre_medicamento'], mov['tipo'], mov['cantidad'], mov['motivo'], mov['fecha']])
        print(f"✅ Datos extraídos guardados en {archivo_salida} para verificación.")

    # ==================== EXTRACCIÓN PRINCIPAL ====================
    def extraer_datos(self, nombre_hoja='CONSOLIDADO', generar_log=False, archivo_log="procesamiento_log.txt"):
        """Extrae todos los datos del Excel desde una hoja consolidada con columna TIPO"""
        # Abrir archivo de log si se solicita
        if generar_log:
            self.log_file = open(archivo_log, 'w', encoding='utf-8')
            self._log(f"Inicio de procesamiento - {datetime.now()}", "INICIO")
            self._log(f"Archivo Excel: {self.archivo_excel}")
            self._log(f"Hoja: {nombre_hoja}")

        self.omitidos = []  # Reiniciar lista de omitidos

        print("\n📂 Cargando archivo Excel...")
        self.wb = load_workbook(self.archivo_excel, data_only=True)

        if nombre_hoja not in self.wb.sheetnames:
            print(f"❌ No se encontró la hoja '{nombre_hoja}'")
            print(f"   Hojas disponibles: {self.wb.sheetnames}")
            if self.log_file:
                self.log_file.close()
            return False

        sheet_name = nombre_hoja
        print(f"📄 Procesando hoja: '{sheet_name}'")
        df_raw = pd.read_excel(self.archivo_excel, sheet_name=sheet_name, header=None)

        # Buscar encabezados
        header_row = None
        for idx in range(min(10, len(df_raw))):
            row = df_raw.iloc[idx]
            if pd.notna(row[0]) and str(row[0]).strip().upper() == 'MEDICAMENTO':
                header_row = idx
                break

        if header_row is None:
            print("❌ No se encontró la fila de encabezados (buscando 'MEDICAMENTO')")
            if self.log_file:
                self.log_file.close()
            return False

        print(f"📍 Encabezados encontrados en fila Excel: {header_row+1}")
        self._log(f"📍 Encabezados encontrados en fila Excel: {header_row+1}")

        encabezados_raw = df_raw.iloc[header_row].tolist()
        encabezados = [str(col).strip() if pd.notna(col) else '' for col in encabezados_raw]

        df = df_raw.iloc[header_row + 1:].reset_index(drop=True)
        df.columns = encabezados

        self._mapear_columnas(df.columns.tolist())

        col_med = self.col_mapping.get('medicamento')
        if col_med is None:
            col_med = df.columns[0]
            print(f"⚠️ Usando primera columna como medicamento: '{col_med}'")
            self._log(f"⚠️ Usando primera columna como medicamento: '{col_med}'")

        ws = self.wb[sheet_name]
        excel_start_row = header_row + 2

        total_filas = len(df)
        print(f"📊 Total de filas en la hoja (después de encabezados): {total_filas}")
        self._log(f"📊 Total de filas en la hoja (después de encabezados): {total_filas}")

        registros_procesados = 0
        registros_omitidos = 0

        print("🔍 Procesando cada fila...\n")
        self._log("🔍 Procesando cada fila...\n")

        for idx, row in df.iterrows():
            fila_excel_num = excel_start_row + idx
            nombre_val = row[col_med]
            if pd.isna(nombre_val):
                motivo = "Nombre vacío o NaN"
                self.omitidos.append([fila_excel_num, motivo, "", ""])
                self._log(f"--- Fila Excel {fila_excel_num} (índice {idx}) ---")
                self._log(f"  ⏭️ OMITIDA: {motivo}")
                registros_omitidos += 1
                continue

            nombre = str(nombre_val).strip()
            if not nombre or nombre.lower() in ('nan', 'none') or len(nombre) < 2:
                motivo = f"Nombre inválido: '{nombre}'"
                self.omitidos.append([fila_excel_num, motivo, nombre, ""])
                self._log(f"--- Fila Excel {fila_excel_num} (índice {idx}) ---")
                self._log(f"  ⏭️ OMITIDA: {motivo}")
                registros_omitidos += 1
                continue

            # Obtener presentación
            col_pres = self.col_mapping.get('presentacion')
            presentacion_raw = ''
            if col_pres and pd.notna(row[col_pres]):
                presentacion_raw = str(row[col_pres]).strip()
                if presentacion_raw.lower() in ('nan', 'none'):
                    presentacion_raw = ''

            # Extraer forma farmacéutica y concentración
            concentracion = self._extraer_concentracion(nombre, presentacion_raw)
            forma_farmaceutica = self._extraer_forma_farmaceutica(presentacion_raw, nombre)
            es_controlado = self._detectar_color_morado(ws, fila_excel_num, columna_letra='A')
            nombre_limpio = self._limpiar_nombre_medicamento(nombre)

            # Ubicación
            col_ubic = self.col_mapping.get('ubicacion')
            ubicacion = None
            if col_ubic and pd.notna(row[col_ubic]):
                ubicacion = str(row[col_ubic]).strip()
                if ubicacion.lower() in ('nan', 'none'):
                    ubicacion = None

            # Caducidad
            col_cad = self.col_mapping.get('caducidad')
            caducidad = None
            if col_cad and pd.notna(row[col_cad]):
                caducidad = self._parsear_fecha_caducidad(row[col_cad])

            # Cantidades
            col_entradas = self.col_mapping.get('entradas')
            entradas = self._extraer_numero(row[col_entradas]) if col_entradas else 0
            col_salidas = self.col_mapping.get('salidas')
            salidas = self._extraer_numero(row[col_salidas]) if col_salidas else 0
            col_stock = self.col_mapping.get('existencias')
            stock_actual = self._extraer_numero(row[col_stock]) if col_stock else 0

            # Observaciones desde Anotaciones
            col_anot = self.col_mapping.get('anotaciones')
            observaciones = ''
            if col_anot and pd.notna(row[col_anot]):
                obs = str(row[col_anot]).strip()
                if obs and obs.lower() not in ('nan', 'none'):
                    observaciones = obs

            # Leer TIPO y DONANTE explícitos
            tipo_registro = ''
            if 'tipo' in self.col_mapping and pd.notna(row[self.col_mapping['tipo']]):
                tipo_registro = str(row[self.col_mapping['tipo']]).strip().upper()

            donante_explicito = None
            if 'donante_col' in self.col_mapping and pd.notna(row[self.col_mapping['donante_col']]):
                donante_explicito = str(row[self.col_mapping['donante_col']]).strip()

            # ==================== LÓGICA DE SUPLEMENTO Y DONACIÓN ====================
            es_donacion = False
            donante = None
            accion_extra = ""

            if tipo_registro == 'SUPLEMENTO':
                if observaciones:
                    observaciones = f"Suplemento. {observaciones}"
                else:
                    observaciones = "Suplemento"
                accion_extra = "→ observaciones actualizadas"
            elif tipo_registro == 'DONACION':
                es_donacion = True
                donante = donante_explicito if donante_explicito else "FARMAMIGO"
                if observaciones:
                    observaciones = f"Donación. {observaciones}"
                else:
                    observaciones = "Donación"
                accion_extra = f"→ donante={donante}"
            else:
                # Si no es DONACION explícita pero en observaciones hay palabras clave de donación
                if not es_donacion and observaciones:
                    if any(p in observaciones.lower() for p in ['donacion', 'donación', 'muestra', 'donado']):
                        es_donacion = True
                        donante = donante_explicito or "Donación sin especificar"
                        accion_extra = "→ detectada por palabras clave"

            # Agregar al catálogo
            self.medicamentos.append({
                'nombre': nombre_limpio,
                'forma_farmaceutica': forma_farmaceutica if forma_farmaceutica else None,
                'concentracion': concentracion if concentracion else None,
                'es_controlado': es_controlado
            })

            # Agregar al inventario (sin lote)
            self.inventario.append({
                'nombre_medicamento': nombre_limpio,
                'ubicacion': ubicacion,
                'caducidad': caducidad,
                'stock': stock_actual,
                'es_donacion': es_donacion,
                'donante': donante,
                'observaciones': observaciones if observaciones else None
            })

            # Generar movimientos
            if entradas > 0:
                self.movimientos.append({
                    'nombre_medicamento': nombre_limpio,
                    'tipo': 'entrada',
                    'cantidad': entradas,
                    'motivo': f'Registro inicial desde Excel - Entradas: {entradas} unidades',
                    'fecha': date.today()
                })
            if salidas > 0:
                self.movimientos.append({
                    'nombre_medicamento': nombre_limpio,
                    'tipo': 'salida',
                    'cantidad': salidas,
                    'motivo': f'Registro inicial desde Excel - Salidas: {salidas} unidades',
                    'fecha': date.today()
                })

            # Ajuste para conciliar stock
            stock_calculado = entradas - salidas
            if stock_calculado != stock_actual and stock_actual > 0:
                diferencia = stock_actual - stock_calculado
                if diferencia > 0:
                    self.movimientos.append({
                        'nombre_medicamento': nombre_limpio,
                        'tipo': 'entrada',
                        'cantidad': diferencia,
                        'motivo': f'Ajuste para conciliar inventario (+{diferencia} unidades) - Carga inicial',
                        'fecha': date.today()
                    })
                elif diferencia < 0:
                    self.movimientos.append({
                        'nombre_medicamento': nombre_limpio,
                        'tipo': 'salida',
                        'cantidad': abs(diferencia),
                        'motivo': f'Ajuste para conciliar inventario ({abs(diferencia)} unidades) - Carga inicial',
                        'fecha': date.today()
                    })

            registros_procesados += 1
            self._log(f"--- Fila Excel {fila_excel_num} (índice {idx}) ---")
            self._log(f"  ✅ Nombre válido: {nombre}")
            self._log(f"     Presentación: {presentacion_raw if presentacion_raw else '(vacío)'}")
            self._log(f"     Tipo: {tipo_registro if tipo_registro else 'NORMAL'} {accion_extra}")

            if registros_procesados % 50 == 0:
                print(f"  Procesados {registros_procesados} registros hasta fila {fila_excel_num}...")
                self._log(f"  Procesados {registros_procesados} registros hasta fila {fila_excel_num}...")

        # Eliminar duplicados del catálogo
        self._eliminar_duplicados_catalogo()

        mov_entradas = sum(1 for m in self.movimientos if m['tipo'] == 'entrada')
        mov_salidas = sum(1 for m in self.movimientos if m['tipo'] == 'salida')

        print(f"\n✅ Procesamiento completado:")
        print(f"  - Total filas leídas en hoja: {total_filas}")
        print(f"  - Registros procesados (incluidos): {registros_procesados}")
        print(f"  - Registros omitidos: {registros_omitidos}")
        print(f"  - Medicamentos únicos en catálogo: {len(self.medicamentos)}")
        print(f"  - Registros de inventario: {len(self.inventario)}")
        print(f"  - Movimientos generados: {len(self.movimientos)} (entradas: {mov_entradas}, salidas: {mov_salidas})")

        self._log(f"\n✅ Procesamiento completado:")
        self._log(f"  - Total filas leídas en hoja: {total_filas}")
        self._log(f"  - Registros procesados (incluidos): {registros_procesados}")
        self._log(f"  - Registros omitidos: {registros_omitidos}")
        self._log(f"  - Medicamentos únicos en catálogo: {len(self.medicamentos)}")
        self._log(f"  - Registros de inventario: {len(self.inventario)}")
        self._log(f"  - Movimientos generados: {len(self.movimientos)} (entradas: {mov_entradas}, salidas: {mov_salidas})")

        # Guardar omitidos en CSV si hay
        if self.omitidos:
            with open("omitidos.csv", 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow(["Fila Excel", "Motivo", "Nombre", "Presentación"])
                writer.writerows(self.omitidos)
            print(f"\n⚠️ Se omitieron {len(self.omitidos)} filas. Revisa el archivo 'omitidos.csv' para detalles.")
            self._log(f"⚠️ Se omitieron {len(self.omitidos)} filas. Revisa el archivo 'omitidos.csv' para detalles.")
        else:
            print(f"\n✅ No se omitió ninguna fila.")
            self._log(f"✅ No se omitió ninguna fila.")

        # Cerrar archivo de log
        if self.log_file:
            self._log(f"\n📄 Log detallado guardado en {archivo_log}")
            self.log_file.close()
            self.log_file = None

        return True

    # ==================== INSERCIÓN EN BASE DE DATOS ====================
    def reiniciar_contadores_identidad(self, cursor):
        """Reinicia los contadores de identidad (IDs) a 1"""
        try:
            print("  🔄 Reiniciando contadores de identidad...")
            cursor.execute("DBCC CHECKIDENT ('catalogo_medicamentos', RESEED, 0)")
            print("     ✅ ID de catálogo reiniciado a 1")
            cursor.execute("DBCC CHECKIDENT ('inventario_farmacia', RESEED, 0)")
            print("     ✅ ID de inventario reiniciado a 1")
            cursor.execute("DBCC CHECKIDENT ('movimientos_farmacia', RESEED, 0)")
            print("     ✅ ID de movimientos reiniciado a 1")
        except pyodbc.Error as e:
            print(f"     ⚠️ Error al reiniciar contadores: {e}")

    def insertar_en_base_datos(self, server, database, limpiar_tablas=False, reiniciar_ids=True, username=None, password=None):
        """Inserta los datos directamente en la base de datos (sin columna lote)"""
        print("\n💾 Conectando a la base de datos...")

        if username and password:
            conn_str = f'DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={server};DATABASE={database};UID={username};PWD={password};'
        else:
            conn_str = f'DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={server};DATABASE={database};Trusted_Connection=yes;'

        try:
            conn = pyodbc.connect(conn_str, timeout=30)
            cursor = conn.cursor()
            print("✅ Conexión exitosa!")

            self._obtener_usuario_admin(cursor)

            if limpiar_tablas:
                print("\n🗑️ Limpiando tablas existentes...")
                cursor.execute("DELETE FROM movimientos_farmacia")
                print("  ✅ Movimientos eliminados")
                cursor.execute("DELETE FROM inventario_farmacia")
                print("  ✅ Inventario eliminado")
                cursor.execute("DELETE FROM catalogo_medicamentos")
                print("  ✅ Catálogo eliminado")

                if reiniciar_ids:
                    self.reiniciar_contadores_identidad(cursor)

            # Insertar catálogo
            print("\n📝 Insertando catálogo de medicamentos...")
            catalogos_insertados = 0
            for med in self.medicamentos:
                cursor.execute("""
                    INSERT INTO catalogo_medicamentos (nombre, forma_farmaceutica, concentracion, es_controlado)
                    VALUES (?, ?, ?, ?)
                """, (med['nombre'], med['forma_farmaceutica'], med['concentracion'], med['es_controlado']))
                catalogos_insertados += 1
                if catalogos_insertados % 50 == 0:
                    print(f"  Insertados {catalogos_insertados} medicamentos...")
            print(f"  ✅ {catalogos_insertados} medicamentos insertados")

            # Insertar inventario (sin lote)
            print("\n📦 Insertando inventario...")
            inventario_insertados = 0
            for inv in self.inventario:
                cursor.execute("SELECT id_catalogo FROM catalogo_medicamentos WHERE nombre = ?", (inv['nombre_medicamento'],))
                row = cursor.fetchone()
                if row:
                    id_catalogo = row[0]
                    cursor.execute("""
                        INSERT INTO inventario_farmacia (id_catalogo, ubicacion, caducidad, stock, es_donacion, donante, observaciones)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    """, (id_catalogo, inv['ubicacion'], inv['caducidad'], inv['stock'],
                          inv['es_donacion'], inv['donante'], inv['observaciones']))
                    inventario_insertados += 1
                    if inventario_insertados % 50 == 0:
                        print(f"  Insertados {inventario_insertados} registros de inventario...")
            print(f"  ✅ {inventario_insertados} registros de inventario insertados")

            # Insertar movimientos
            print("\n📋 Insertando movimientos...")
            movimientos_insertados = 0
            mov_entradas = 0
            mov_salidas = 0

            for mov in self.movimientos:
                cursor.execute("""
                    SELECT TOP 1 i.id_inventario
                    FROM inventario_farmacia i
                    INNER JOIN catalogo_medicamentos c ON i.id_catalogo = c.id_catalogo
                    WHERE c.nombre = ?
                """, (mov['nombre_medicamento'],))
                row = cursor.fetchone()
                if row:
                    id_inventario = row[0]
                    if mov['tipo'] == 'entrada':
                        quien_recibe = self.usuario_admin_nombre
                        quien_retira = None
                    else:
                        quien_recibe = None
                        quien_retira = self.usuario_admin_nombre

                    cursor.execute("""
                        INSERT INTO movimientos_farmacia (id_inventario, tipo, cantidad, quien_recibe, quien_retira, motivo, fecha)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    """, (id_inventario, mov['tipo'], mov['cantidad'], quien_recibe, quien_retira, mov['motivo'], mov['fecha']))
                    movimientos_insertados += 1
                    if mov['tipo'] == 'entrada':
                        mov_entradas += 1
                    else:
                        mov_salidas += 1

                    if movimientos_insertados % 100 == 0:
                        print(f"  Insertados {movimientos_insertados} movimientos...")

            print(f"  ✅ {movimientos_insertados} movimientos insertados")
            print(f"     · Entradas: {mov_entradas}")
            print(f"     · Salidas: {mov_salidas}")
            print(f"     · Asignados al usuario: {self.usuario_admin_nombre}")

            conn.commit()
            cursor.close()
            conn.close()

            print("\n" + "=" * 70)
            print("✅ ¡ÉXITO! Todos los datos fueron insertados correctamente")
            print("=" * 70)
            return True

        except pyodbc.Error as e:
            print(f"\n❌ Error de base de datos: {e}")
            return False
        except Exception as e:
            print(f"\n❌ Error inesperado: {e}")
            import traceback
            traceback.print_exc()
            return False


# ==================== MAIN ====================
def main():
    print("=" * 70)
    print("🏥 EXTRACCIÓN E INSERCIÓN DE DATOS PARA FARMACIA")
    print("=" * 70)

    print("\n📁 CONFIGURACIÓN DEL ARCHIVO")
    archivo_default = 'INVENTARIO ACTUALIZADO 21-11.xlsx'
    archivo = input(f"Nombre del archivo Excel [{archivo_default}]: ").strip()
    if not archivo:
        archivo = archivo_default

    if not os.path.exists(archivo):
        print(f"❌ Error: No se encuentra el archivo '{archivo}'")
        return

    # Nombre de la hoja consolidada
    hoja_default = "CONSOLIDADO"
    hoja = input(f"Nombre de la hoja a procesar (con columna TIPO) [{hoja_default}]: ").strip()
    if not hoja:
        hoja = hoja_default

    # Preguntar si se quiere generar log detallado
    generar_log = input("¿Generar log detallado del procesamiento? (s/n): ").strip().lower() == 's'

    print("\n" + "=" * 70)
    print("📊 EXTRACCIÓN DE DATOS")
    print("=" * 70)

    extractor = ExtraerDatosFarmacia(archivo)

    if not extractor.extraer_datos(nombre_hoja=hoja, generar_log=generar_log, archivo_log="procesamiento_log.txt"):
        print("❌ Error al extraer datos")
        return

    # Verificación opcional
    verificar = input("\n¿Deseas exportar los datos extraídos a CSV para verificar? (s/n): ").strip().lower()
    if verificar == 's':
        csv_name = input("Nombre del archivo CSV [verificacion_extraccion.csv]: ").strip()
        if not csv_name:
            csv_name = "verificacion_extraccion.csv"
        extractor.exportar_a_csv_para_verificacion(csv_name)

    print("\n" + "=" * 70)
    print("📋 RESUMEN DE DATOS EXTRAÍDOS")
    print("=" * 70)
    print(f"  Medicamentos únicos: {len(extractor.medicamentos)}")
    print(f"  Registros de inventario: {len(extractor.inventario)}")
    mov_entradas = sum(1 for m in extractor.movimientos if m['tipo'] == 'entrada')
    mov_salidas = sum(1 for m in extractor.movimientos if m['tipo'] == 'salida')
    print(f"  Movimientos a insertar: {len(extractor.movimientos)}")
    print(f"    · Entradas: {mov_entradas}")
    print(f"    · Salidas: {mov_salidas}")

    print("\n" + "=" * 70)
    print("💾 CONFIGURACIÓN DE BASE DE DATOS")
    print("=" * 70)

    respuesta = input("\n¿Deseas insertar los datos en la base de datos? (s/n): ").strip().lower()
    if respuesta != 's':
        print("❌ Operación cancelada por el usuario")
        return

    print("\n🔧 CONFIGURACIÓN DE CONEXIÓN SQL SERVER")
    server = input("Servidor (ej: localhost\\SQLEXPRESS o (local)): ").strip()
    database = input("Nombre de la base de datos: ").strip()

    if not server or not database:
        print("❌ Servidor y base de datos son obligatorios")
        return

    print("\n🔐 AUTENTICACIÓN")
    print("1. Autenticación de Windows (Trusted Connection)")
    print("2. Autenticación SQL Server")
    auth_opcion = input("Selecciona una opción (1 o 2): ").strip()

    username = None
    password = None

    if auth_opcion == '2':
        username = input("Usuario: ").strip()
        password = input("Contraseña: ").strip()
        if not username or not password:
            print("❌ Usuario y contraseña son obligatorios")
            return
    elif auth_opcion != '1':
        print("❌ Opción no válida, usando autenticación de Windows")

    print("\n🗑️ OPCIÓN DE LIMPIEZA")
    limpiar = input("¿Deseas limpiar las tablas existentes antes de insertar? (s/n): ").strip().lower() == 's'

    reiniciar_ids = False
    if limpiar:
        reiniciar = input("¿Deseas reiniciar los contadores de ID (que empiecen desde 1)? (s/n): ").strip().lower()
        reiniciar_ids = reiniciar == 's'

    print("\n" + "=" * 70)
    print("📋 RESUMEN FINAL")
    print("=" * 70)
    print(f"  Servidor: {server}")
    print(f"  Base de datos: {database}")
    print(f"  Limpiar tablas: {'Sí' if limpiar else 'No'}")
    if limpiar:
        print(f"  Reiniciar IDs: {'Sí' if reiniciar_ids else 'No'}")
    print(f"  Medicamentos a insertar: {len(extractor.medicamentos)}")
    print(f"  Inventario a insertar: {len(extractor.inventario)}")
    print(f"  Movimientos a insertar: {len(extractor.movimientos)}")
    print(f"    · Entradas: {mov_entradas}")
    print(f"    · Salidas: {mov_salidas}")
    print(f"  Todos los movimientos se asignarán al usuario administrador del sistema")

    confirmar = input("\n¿Confirmas la inserción? (s/n): ").strip().lower()
    if confirmar != 's':
        print("❌ Operación cancelada")
        return

    success = extractor.insertar_en_base_datos(server, database, limpiar, reiniciar_ids, username, password)

    if not success:
        print("\n❌ No se pudo completar la inserción")
        return

    print("\n✨ ¡Proceso completado exitosamente!")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n❌ Proceso interrumpido por el usuario")
    except Exception as e:
        print(f"\n❌ Error inesperado: {e}")
        import traceback
        traceback.print_exc()