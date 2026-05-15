import customtkinter as ctk
from tkinter import messagebox, ttk, filedialog
from datetime import datetime, date
import calendar
from assets.styles.themes import AppTheme
from models.medicamento import Medicamento
from models.movimiento import Movimiento
from config.database import Database
from utils.icon_manager import IconManager

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class SistemaView(ctk.CTk):

    def __init__(self, usuario, area_id, area_nombre):
        super().__init__()
        IconManager.aplicar_icono(self)

        self.usuario = usuario
        self.area_id = area_id
        self.area_nombre = area_nombre
        self.id_medicamento = None
        self.id_movimiento_editando = None
        self.fecha_desde = date(date.today().year, 1, 1)
        self.fecha_hasta = date.today()
        self._debug = True
        self.fullscreen_flag = False

        self.medicamento_model = Medicamento
        self.movimiento_model = Movimiento

        self.title(f"MediKinv - {area_nombre}")

        # Configuración responsiva
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()

        # Usar casi toda la pantalla
        window_width = screen_width - 40
        window_height = screen_height - 80

        # Tamaño mínimo
        min_width = 800
        min_height = 600

        final_width = max(window_width, min_width)
        final_height = max(window_height, min_height)

        # Maximizar por defecto
        self.geometry(f"{final_width}x{final_height}")
        self.minsize(min_width, min_height)

        # Intentar maximizar
        try:
            self.state('zoomed')
        except:
            pass

        AppTheme.aplicar_tema()
        self.configurar_estilos()
        self.crear_interfaz()
        self.after(200, self.cargar_datos_iniciales)

        self.bind('<Escape>', self.toggle_fullscreen)
        self.bind('<Configure>', self.on_window_resize)

    def debug_print(self, msg):
        if self._debug:
            print(f"[DEBUG {self.area_nombre}] {msg}")

    def on_window_resize(self, event=None):
        if event and event.widget == self:
            self.after(100, self.ajustar_tablas)

    def ajustar_tablas(self):
        """Ajusta dinámicamente el tamaño de las tablas"""
        if hasattr(self, 'frame_tabla_inventario'):
            altura = self.winfo_height() - 350
            altura_filas = max(15, altura // 30)

            # Reconfigurar altura de las tablas
            if hasattr(self, 'tabla_inventario'):
                self.tabla_inventario.configure(height=altura_filas)

            if hasattr(self, 'tabla_mov'):
                self.tabla_mov.configure(height=altura_filas)

            # Ajustar anchos de columnas
            ancho = self.winfo_width()
            self.ajustar_columnas_inventario(ancho)
            self.ajustar_columnas_movimientos(ancho)

    def ajustar_columnas_inventario(self, ancho):
        if not hasattr(self, 'tabla_inventario'):
            return

        # Calcular anchos proporcionales
        if ancho < 1000:
            self.tabla_inventario.column("ID", width=40)
            self.tabla_inventario.column("Nombre", width=120)
            self.tabla_inventario.column("Presentacion", width=100)
            self.tabla_inventario.column("Ubicacion", width=80)
            self.tabla_inventario.column("Stock", width=50)
            self.tabla_inventario.column("Caducidad", width=85)
            self.tabla_inventario.column("Donación", width=60)
            self.tabla_inventario.column("Donante", width=100)
            self.tabla_inventario.column("Observaciones", width=120)
        elif ancho < 1200:
            self.tabla_inventario.column("ID", width=50)
            self.tabla_inventario.column("Nombre", width=160)
            self.tabla_inventario.column("Presentacion", width=130)
            self.tabla_inventario.column("Ubicacion", width=110)
            self.tabla_inventario.column("Stock", width=60)
            self.tabla_inventario.column("Caducidad", width=95)
            self.tabla_inventario.column("Donación", width=70)
            self.tabla_inventario.column("Donante", width=130)
            self.tabla_inventario.column("Observaciones", width=160)
        else:
            self.tabla_inventario.column("ID", width=60)
            self.tabla_inventario.column("Nombre", width=220)
            self.tabla_inventario.column("Presentacion", width=170)
            self.tabla_inventario.column("Ubicacion", width=140)
            self.tabla_inventario.column("Stock", width=70)
            self.tabla_inventario.column("Caducidad", width=105)
            self.tabla_inventario.column("Donación", width=80)
            self.tabla_inventario.column("Donante", width=160)
            self.tabla_inventario.column("Observaciones", width=200)

    def ajustar_columnas_movimientos(self, ancho):
        if not hasattr(self, 'tabla_mov'):
            return

        if ancho < 1000:
            self.tabla_mov.column("ID", width=40)
            self.tabla_mov.column("Medicamento", width=120)
            self.tabla_mov.column("Tipo", width=70)
            self.tabla_mov.column("Cantidad", width=60)
            self.tabla_mov.column("Quien Recibe", width=100)
            self.tabla_mov.column("Quien Retira", width=100)
            self.tabla_mov.column("Motivo", width=150)
            self.tabla_mov.column("Fecha", width=85)
        elif ancho < 1200:
            self.tabla_mov.column("ID", width=50)
            self.tabla_mov.column("Medicamento", width=160)
            self.tabla_mov.column("Tipo", width=85)
            self.tabla_mov.column("Cantidad", width=70)
            self.tabla_mov.column("Quien Recibe", width=130)
            self.tabla_mov.column("Quien Retira", width=130)
            self.tabla_mov.column("Motivo", width=200)
            self.tabla_mov.column("Fecha", width=95)
        else:
            self.tabla_mov.column("ID", width=60)
            self.tabla_mov.column("Medicamento", width=200)
            self.tabla_mov.column("Tipo", width=100)
            self.tabla_mov.column("Cantidad", width=80)
            self.tabla_mov.column("Quien Recibe", width=160)
            self.tabla_mov.column("Quien Retira", width=160)
            self.tabla_mov.column("Motivo", width=250)
            self.tabla_mov.column("Fecha", width=105)

    def cargar_datos_iniciales(self):
        self.cargar_inventario()
        self.cargar_combo_medicamentos_filtro()
        self.cargar_combo_ubicaciones()
        self.cargar_combo_medicamentos_mov()
        self.cargar_historial()
        self.ajustar_tablas()

    def configurar_estilos(self):
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("Inventory.Treeview", font=("Segoe UI", 10), rowheight=32,
                        background="white", fieldbackground="white", foreground=AppTheme.TEXT_PRIMARY)
        style.configure("Inventory.Treeview.Heading", font=("Segoe UI", 10, "bold"),
                        background=AppTheme.PRIMARY, foreground=AppTheme.TEXT_ON_PRIMARY, relief="flat")
        style.configure("Movements.Treeview", font=("Segoe UI", 10), rowheight=32,
                        background="white", fieldbackground="white", foreground=AppTheme.TEXT_PRIMARY)
        style.configure("Movements.Treeview.Heading", font=("Segoe UI", 10, "bold"),
                        background=AppTheme.PRIMARY, foreground=AppTheme.TEXT_ON_PRIMARY, relief="flat")

    def crear_interfaz(self):
        self.frame_principal = ctk.CTkFrame(self, fg_color=AppTheme.BACKGROUND, corner_radius=0)
        self.frame_principal.pack(fill="both", expand=True)
        self.crear_header()

        # Tabview principal
        self.tab_view = ctk.CTkTabview(self.frame_principal, fg_color=AppTheme.SURFACE, corner_radius=10)
        self.tab_view.pack(fill="both", expand=True, padx=10, pady=10)

        self.tab_inventario = self.tab_view.add("📦 INVENTARIO")
        self.tab_movimientos = self.tab_view.add("📋 MOVIMIENTOS")

        self.crear_pestana_inventario()
        self.crear_pestana_movimientos()

    def crear_header(self):
        header = ctk.CTkFrame(self.frame_principal, fg_color=AppTheme.PRIMARY, height=60, corner_radius=0)
        header.pack(fill="x")
        header.pack_propagate(False)

        icono = "🏥" if self.area_id == 1 else "🦷"
        titulo = ctk.CTkLabel(header, text=f"{icono} MediKinv - {self.area_nombre}",
                              font=("Segoe UI", 18, "bold"), text_color=AppTheme.TEXT_ON_PRIMARY)
        titulo.pack(side="left", padx=20)

        frame_usuario = ctk.CTkFrame(header, fg_color="transparent")
        frame_usuario.pack(side="right", padx=20)

        label_usuario = ctk.CTkLabel(frame_usuario, text=f"👤 {self.usuario.nombre_usuario}",
                                     font=("Segoe UI", 11), text_color=AppTheme.TEXT_ON_PRIMARY)
        label_usuario.pack(side="left", padx=10)

        btn_fullscreen = ctk.CTkButton(frame_usuario, text="⛶", width=40, height=32,
                                       font=("Segoe UI", 14), fg_color="transparent",
                                       hover_color=AppTheme.PRIMARY_LIGHT, text_color=AppTheme.TEXT_ON_PRIMARY,
                                       corner_radius=8, command=self.toggle_fullscreen)
        btn_fullscreen.pack(side="left", padx=5)

        btn_volver = ctk.CTkButton(frame_usuario, text="← Áreas", width=80, height=32,
                                   font=("Segoe UI", 11), fg_color="transparent",
                                   hover_color=AppTheme.PRIMARY_LIGHT, text_color=AppTheme.TEXT_ON_PRIMARY,
                                   corner_radius=8, command=self.volver_areas)
        btn_volver.pack(side="left", padx=5)

        btn_cerrar = ctk.CTkButton(frame_usuario, text="🚪 Salir", width=80, height=32,
                                   font=("Segoe UI", 11), fg_color=AppTheme.PRIMARY_LIGHT,
                                   hover_color=AppTheme.PRIMARY_DARK, text_color=AppTheme.TEXT_ON_PRIMARY,
                                   corner_radius=8, command=self.cerrar_sesion)
        btn_cerrar.pack(side="left", padx=5)

    def toggle_fullscreen(self, event=None):
        self.fullscreen_flag = not self.fullscreen_flag
        self.attributes('-fullscreen', self.fullscreen_flag)
        if not self.fullscreen_flag:
            self.state('zoomed')
        self.after(100, self.ajustar_tablas)

    def volver_areas(self):
        try:
            for after_id in self.tk.eval('after info').split():
                try:
                    self.after_cancel(int(after_id))
                except:
                    pass
        except:
            pass
        from views.area_select_view import AreaSelectView
        area_view = AreaSelectView(self.usuario)
        self.destroy()
        area_view.mainloop()

    def cerrar_sesion(self):
        try:
            for after_id in self.tk.eval('after info').split():
                try:
                    self.after_cancel(int(after_id))
                except:
                    pass
        except:
            pass
        from views.login_view import LoginView
        login = LoginView()
        self.destroy()
        login.mainloop()

    # ---------- PESTAÑA INVENTARIO ----------
    def crear_pestana_inventario(self):
        # Frame principal
        main_frame = ctk.CTkFrame(self.tab_inventario, fg_color="transparent")
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)

        # Frame de botones superiores
        botones_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        botones_frame.pack(fill="x", pady=(0, 10))

        btn_nuevo = ctk.CTkButton(botones_frame, text="➕ NUEVO MEDICAMENTO", width=200, height=40,
                                  font=("Segoe UI", 12, "bold"), fg_color=AppTheme.PRIMARY,
                                  hover_color=AppTheme.PRIMARY_DARK, corner_radius=10,
                                  command=self.abrir_formulario_medicamento)
        btn_nuevo.pack(side="left", padx=5)

        btn_nuevo_mov = ctk.CTkButton(botones_frame, text="📊 REGISTRAR MOVIMIENTO", width=200, height=40,
                                      font=("Segoe UI", 12, "bold"), fg_color="#2c7a4d",
                                      hover_color="#1e5a38", corner_radius=10,
                                      command=self.abrir_formulario_movimiento_rapido)
        btn_nuevo_mov.pack(side="left", padx=5)

        # Filtros
        self.crear_filtros_inventario(main_frame)

        # Leyenda de colores
        self.crear_leyenda_colores(main_frame)

        # Tabla
        self.crear_tabla_inventario(main_frame)

    def crear_leyenda_colores(self, parent):
        frame_leyenda = ctk.CTkFrame(parent, fg_color=AppTheme.SURFACE, corner_radius=10)
        frame_leyenda.pack(fill="x", pady=(0, 10), padx=5)

        colores = [
            ("#D4A837", "Sin stock"),
            ("#C75C5C", "Caducado"),
            ("#D97A3A", "Caduca mes"),
            ("#5A9E6F", "Nuevo (<7 días)"),
            ("#8B6B9E", "Controlado"),
            ("#5B7BA5", "Normal"),
        ]

        for color, texto in colores:
            frame_item = ctk.CTkFrame(frame_leyenda, fg_color="transparent")
            frame_item.pack(side="left", padx=10, pady=5)

            cuadro = ctk.CTkFrame(frame_item, width=16, height=16, fg_color=color, corner_radius=4)
            cuadro.pack(side="left")
            cuadro.pack_propagate(False)

            ctk.CTkLabel(frame_item, text=texto, font=("Segoe UI", 10),
                         text_color=AppTheme.TEXT_SECONDARY).pack(side="left", padx=(5, 0))

    def crear_filtros_inventario(self, parent):
        frame_filtros = ctk.CTkFrame(parent, fg_color=AppTheme.SURFACE, corner_radius=10)
        frame_filtros.pack(fill="x", pady=(0, 10), padx=5)

        # Fila 1 - Búsqueda
        frame_buscar = ctk.CTkFrame(frame_filtros, fg_color="transparent")
        frame_buscar.pack(fill="x", padx=10, pady=10)

        self.txt_buscar = ctk.CTkEntry(frame_buscar, placeholder_text="🔍 Buscar por nombre o descripción...",
                                       height=38, font=("Segoe UI", 12), corner_radius=10)
        self.txt_buscar.pack(side="left", fill="x", expand=True, padx=(0, 10))
        self.txt_buscar.bind("<KeyRelease>", self.buscar_medicamento)

        btn_actualizar = ctk.CTkButton(frame_buscar, text="🔄 Actualizar", width=120, height=38,
                                       font=("Segoe UI", 11), fg_color=AppTheme.PRIMARY,
                                       hover_color=AppTheme.PRIMARY_DARK, corner_radius=10,
                                       command=self.cargar_inventario)
        btn_actualizar.pack(side="right", padx=5)

        btn_exportar = ctk.CTkButton(frame_buscar, text="📎 Exportar", width=100, height=38,
                                     font=("Segoe UI", 11), fg_color="#2c7a4d",
                                     hover_color="#1e5a38", corner_radius=10,
                                     command=self.exportar_excel_inventario)
        btn_exportar.pack(side="right", padx=5)

        # Fila 2 - Filtros rápidos
        frame_filtros2 = ctk.CTkFrame(frame_filtros, fg_color="transparent")
        frame_filtros2.pack(fill="x", padx=10, pady=(0, 10))

        ctk.CTkLabel(frame_filtros2, text="Presentación:", font=("Segoe UI", 11)).pack(side="left", padx=5)
        self.txt_presentacion_filtro = ctk.CTkEntry(frame_filtros2, width=130, height=35,
                                                    corner_radius=10, placeholder_text="Ej: Tableta")
        self.txt_presentacion_filtro.pack(side="left", padx=5)
        self.txt_presentacion_filtro.bind("<KeyRelease>", self.buscar_medicamento)

        ctk.CTkLabel(frame_filtros2, text="Ubicación:", font=("Segoe UI", 11)).pack(side="left", padx=10)
        self.combo_ubicacion_filtro = ctk.CTkComboBox(frame_filtros2, values=["Todas"], width=130,
                                                      height=35, corner_radius=10, state="readonly")
        self.combo_ubicacion_filtro.pack(side="left", padx=5)
        self.combo_ubicacion_filtro.bind("<<ComboboxSelected>>", self.buscar_medicamento)

        ctk.CTkLabel(frame_filtros2, text="Stock:", font=("Segoe UI", 11)).pack(side="left", padx=10)
        self.combo_stock = ctk.CTkComboBox(frame_filtros2,
                                           values=["Todos", "Sin stock", "Bajo (1-29)", "Normal (30-100)",
                                                   "Alto (>100)"],
                                           width=130, height=35, corner_radius=10, state="readonly")
        self.combo_stock.pack(side="left", padx=5)
        self.combo_stock.bind("<<ComboboxSelected>>", self.buscar_medicamento)
        self.combo_stock.set("Todos")

        btn_limpiar = ctk.CTkButton(frame_filtros2, text="🗑 Limpiar filtros", width=120, height=35,
                                    font=("Segoe UI", 11), fg_color="#4a5568", hover_color="#2d3748",
                                    corner_radius=10, command=self.limpiar_filtros_inventario)
        btn_limpiar.pack(side="right", padx=10)

    def limpiar_filtros_inventario(self):
        self.txt_buscar.delete(0, "end")
        self.txt_presentacion_filtro.delete(0, "end")
        self.combo_ubicacion_filtro.set("Todas")
        self.combo_stock.set("Todos")
        self.cargar_inventario()

    def cargar_combo_ubicaciones(self):
        try:
            tabla_inventario = "inventario_farmacia" if self.area_id == 1 else "inventario_odontologia"
            query = f"SELECT DISTINCT ubicacion FROM {tabla_inventario} WHERE ubicacion IS NOT NULL AND ubicacion != ''"
            result = Database.execute_query(query)
            ubicaciones = sorted([row[0] for row in result]) if result else []
            valores = ["Todas"] + ubicaciones
            self.combo_ubicacion_filtro.configure(values=valores)
            if self.combo_ubicacion_filtro.get() not in valores:
                self.combo_ubicacion_filtro.set("Todas")
        except Exception as e:
            self.debug_print(f"Error cargando ubicaciones: {e}")

    def cargar_inventario(self, event=None):
        try:
            for item in self.tabla_inventario.get_children():
                self.tabla_inventario.delete(item)

            busqueda = self.txt_buscar.get().strip()
            filtro_presentacion = self.txt_presentacion_filtro.get().strip().lower()
            filtro_ubicacion = self.combo_ubicacion_filtro.get()
            filtro_stock = self.combo_stock.get()

            tabla_inventario = "inventario_farmacia" if self.area_id == 1 else "inventario_odontologia"
            tabla_movimientos = "movimientos_farmacia" if self.area_id == 1 else "movimientos_odontologia"

            query = f"""
                SELECT 
                    i.id_inventario,
                    c.nombre,
                    COALESCE(c.forma_farmaceutica, '') || ' ' || COALESCE(c.concentracion, '') as presentacion,
                    i.ubicacion,
                    i.stock,
                    i.caducidad,
                    i.es_donacion,
                    i.donante,
                    i.observaciones,
                    c.es_controlado,
                    (SELECT MIN(fecha) FROM {tabla_movimientos} WHERE id_inventario = i.id_inventario) as fecha_primer_movimiento
                FROM {tabla_inventario} i
                JOIN catalogo_medicamentos c ON i.id_catalogo = c.id_catalogo
                WHERE 1=1
            """
            params = []

            if busqueda:
                query += " AND (c.nombre LIKE ? OR i.observaciones LIKE ?)"
                params.append(f"%{busqueda}%")
                params.append(f"%{busqueda}%")

            query += " ORDER BY i.id_inventario"

            result = Database.execute_query(query, tuple(params) if params else None)
            if not result:
                self.tabla_inventario.insert("", "end",
                                             values=("", "", "", "", "", "", "", "", "No hay medicamentos"))
                return

            fecha_hoy = date.today()
            mes_actual = fecha_hoy.month
            anio_actual = fecha_hoy.year
            registros_mostrados = 0

            for row in result:
                id_med = row[0]
                nombre = row[1]
                presentacion = (row[2] or "---").strip()
                ubicacion = row[3] or ""
                stock = row[4]
                caducidad = row[5]
                es_donacion = row[6]
                donante = row[7] or ""
                observaciones = row[8] or ""
                es_controlado = row[9]
                fecha_primer_movimiento = row[10]

                if filtro_presentacion and filtro_presentacion not in presentacion.lower():
                    continue

                if filtro_ubicacion != "Todas" and ubicacion != filtro_ubicacion:
                    continue

                if stock == 0:
                    estado_stock = "Sin stock"
                elif 1 <= stock <= 29:
                    estado_stock = "Bajo (1-29)"
                elif 30 <= stock <= 100:
                    estado_stock = "Normal (30-100)"
                else:
                    estado_stock = "Alto (>100)"

                if filtro_stock != "Todos" and estado_stock != filtro_stock:
                    continue

                donacion_str = "✓" if es_donacion else "✗"
                caducidad_str = caducidad.strftime("%d/%m/%Y") if caducidad else "---"

                if len(observaciones) > 40:
                    observaciones = observaciones[:37] + "..."

                # Determinar color
                if stock == 0:
                    tag = 'sin_stock'
                elif caducidad and caducidad < fecha_hoy:
                    tag = 'caducado'
                elif caducidad and caducidad.month == mes_actual and caducidad.year == anio_actual:
                    tag = 'caduca_mes'
                elif fecha_primer_movimiento and (fecha_hoy - fecha_primer_movimiento).days <= 7:
                    tag = 'nuevo'
                elif es_controlado:
                    tag = 'controlado'
                else:
                    tag = 'normal'

                self.tabla_inventario.insert("", "end", values=(
                    id_med, nombre, presentacion, ubicacion, stock,
                    caducidad_str, donacion_str, donante, observaciones
                ), tags=(tag,))
                registros_mostrados += 1

            if registros_mostrados == 0:
                self.tabla_inventario.insert("", "end",
                                             values=("", "", "", "", "", "", "", "",
                                                     "No hay medicamentos con esos filtros"))

            # Configurar colores
            self.tabla_inventario.tag_configure('sin_stock', background='#D4A837', foreground='#FFFFFF')
            self.tabla_inventario.tag_configure('caducado', background='#C75C5C', foreground='#FFFFFF')
            self.tabla_inventario.tag_configure('caduca_mes', background='#D97A3A', foreground='#FFFFFF')
            self.tabla_inventario.tag_configure('nuevo', background='#5A9E6F', foreground='#FFFFFF')
            self.tabla_inventario.tag_configure('controlado', background='#8B6B9E', foreground='#FFFFFF')
            self.tabla_inventario.tag_configure('normal', background='#5B7BA5', foreground='#FFFFFF')

        except Exception as e:
            self.debug_print(f"Error al cargar inventario: {e}")
            messagebox.showerror("Error", f"No se pudo cargar el inventario:\n{e}")

    def buscar_medicamento(self, event):
        self.cargar_inventario()

    def crear_tabla_inventario(self, parent):
        self.frame_tabla_inventario = ctk.CTkFrame(parent, fg_color="transparent")
        self.frame_tabla_inventario.pack(fill="both", expand=True, padx=5)

        scroll_y = ttk.Scrollbar(self.frame_tabla_inventario, orient="vertical")
        scroll_x = ttk.Scrollbar(self.frame_tabla_inventario, orient="horizontal")

        self.tabla_inventario = ttk.Treeview(
            self.frame_tabla_inventario,
            columns=("ID", "Nombre", "Presentacion", "Ubicacion", "Stock", "Caducidad", "Donación", "Donante",
                     "Observaciones"),
            show="headings",
            yscrollcommand=scroll_y.set,
            xscrollcommand=scroll_x.set,
            style="Inventory.Treeview",
            height=20
        )

        scroll_y.config(command=self.tabla_inventario.yview)
        scroll_x.config(command=self.tabla_inventario.xview)

        columnas = [
            ("ID", 50, "center"),
            ("Nombre", 180, "w"),
            ("Presentacion", 150, "w"),
            ("Ubicacion", 120, "w"),
            ("Stock", 60, "center"),
            ("Caducidad", 95, "center"),
            ("Donación", 60, "center"),
            ("Donante", 140, "w"),
            ("Observaciones", 180, "w")
        ]

        for col, width, anchor in columnas:
            self.tabla_inventario.heading(col, text=col)
            self.tabla_inventario.column(col, width=width, anchor=anchor)

        self.tabla_inventario.grid(row=0, column=0, sticky="nsew")
        scroll_y.grid(row=0, column=1, sticky="ns")
        scroll_x.grid(row=1, column=0, sticky="ew")

        self.frame_tabla_inventario.grid_rowconfigure(0, weight=1)
        self.frame_tabla_inventario.grid_columnconfigure(0, weight=1)

        # Doble click para editar
        self.tabla_inventario.bind("<Double-1>", self.on_medicamento_doble_click)

    def on_medicamento_doble_click(self, event):
        seleccion = self.tabla_inventario.selection()
        if seleccion:
            valores = self.tabla_inventario.item(seleccion[0])['values']
            if valores and valores[0]:
                self.abrir_formulario_medicamento(medicamento_id=valores[0])

    # ---------- FORMULARIO MEDICAMENTO (VENTANA EMERGENTE) ----------
    def abrir_formulario_medicamento(self, medicamento_id=None):
        dialog = ctk.CTkToplevel(self)
        dialog.title("Medicamento" if not medicamento_id else "Editar Medicamento")
        dialog.geometry("600x700")
        dialog.resizable(False, False)
        dialog.grab_set()

        # Centrar
        dialog.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() // 2) - 300
        y = self.winfo_y() + (self.winfo_height() // 2) - 350
        dialog.geometry(f"+{x}+{y}")

        scroll = ctk.CTkScrollableFrame(dialog, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=20, pady=20)

        # Título
        titulo = ctk.CTkLabel(scroll, text="AGREGAR MEDICAMENTO" if not medicamento_id else "EDITAR MEDICAMENTO",
                              font=("Segoe UI", 18, "bold"), text_color=AppTheme.PRIMARY)
        titulo.pack(pady=(0, 15))

        # Buscador de catálogo
        ctk.CTkLabel(scroll, text="Medicamento", font=("Segoe UI", 12, "bold")).pack(anchor="w", pady=(5, 2))
        search_entry = ctk.CTkEntry(scroll, placeholder_text="Escriba para buscar...", height=38,
                                    font=("Segoe UI", 11), corner_radius=10)
        search_entry.pack(fill="x", pady=(0, 5))

        combo_catalogo = ctk.CTkComboBox(scroll, values=[], height=38,
                                         font=("Segoe UI", 11), corner_radius=10, state="readonly")
        combo_catalogo.pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(scroll, text="Presentación", font=("Segoe UI", 12, "bold")).pack(anchor="w", pady=(5, 2))
        txt_presentacion = ctk.CTkEntry(scroll, height=38, font=("Segoe UI", 11),
                                        corner_radius=10, state="readonly")
        txt_presentacion.pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(scroll, text="Ubicación", font=("Segoe UI", 12, "bold")).pack(anchor="w", pady=(5, 2))
        txt_ubicacion = ctk.CTkEntry(scroll, height=38, font=("Segoe UI", 11), corner_radius=10)
        txt_ubicacion.pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(scroll, text="Stock", font=("Segoe UI", 12, "bold")).pack(anchor="w", pady=(5, 2))
        txt_stock = ctk.CTkEntry(scroll, height=38, font=("Segoe UI", 11), corner_radius=10)
        txt_stock.pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(scroll, text="Fecha de Caducidad", font=("Segoe UI", 12, "bold")).pack(anchor="w", pady=(5, 2))
        frame_fecha = ctk.CTkFrame(scroll, fg_color="transparent")
        frame_fecha.pack(fill="x", pady=(0, 10))
        txt_caducidad = ctk.CTkEntry(frame_fecha, placeholder_text="DD/MM/AAAA", height=38,
                                     font=("Segoe UI", 11), corner_radius=10)
        txt_caducidad.pack(side="left", fill="x", expand=True, padx=(0, 5))

        btn_cal = ctk.CTkButton(frame_fecha, text="📅", width=50, height=38, font=("Segoe UI", 14),
                                fg_color=AppTheme.PRIMARY, hover_color=AppTheme.PRIMARY_DARK,
                                corner_radius=10)
        btn_cal.pack(side="right")

        # Donación
        es_donacion = ctk.BooleanVar(value=False)
        check_donacion = ctk.CTkCheckBox(scroll, text="Este medicamento es una donación",
                                         variable=es_donacion, font=("Segoe UI", 12))
        check_donacion.pack(anchor="w", pady=(5, 5))

        frame_donante = ctk.CTkFrame(scroll, fg_color="transparent")
        ctk.CTkLabel(frame_donante, text="Donante", font=("Segoe UI", 12, "bold")).pack(anchor="w")
        txt_donante = ctk.CTkEntry(frame_donante, height=38, font=("Segoe UI", 11),
                                   corner_radius=10, state="disabled")
        txt_donante.pack(fill="x", pady=(2, 0))
        frame_donante.pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(scroll, text="Observaciones", font=("Segoe UI", 12, "bold")).pack(anchor="w", pady=(5, 2))
        txt_observaciones = ctk.CTkTextbox(scroll, height=80, font=("Segoe UI", 11), corner_radius=10)
        txt_observaciones.pack(fill="x", pady=(0, 10))

        # Cargar datos
        catalogo_items = []
        id_catalogo_seleccionado = None

        def cargar_catalogo(busqueda=""):
            nonlocal catalogo_items
            catalogo = self.medicamento_model.obtener_catalogo(busqueda)
            catalogo_items = []
            valores = []
            for item in catalogo:
                catalogo_items.append({
                    'id_catalogo': item['id_catalogo'],
                    'nombre': item['nombre'],
                    'presentacion': item['presentacion']
                })
                valores.append(f"{item['nombre']} - {item['presentacion']}")
            combo_catalogo.configure(values=valores)
            if valores:
                combo_catalogo.set(valores[0])
                on_catalogo_seleccionado(None)

        def on_catalogo_seleccionado(choice):
            nonlocal id_catalogo_seleccionado
            seleccion = combo_catalogo.get()
            if not seleccion:
                return
            for item in catalogo_items:
                if f"{item['nombre']} - {item['presentacion']}" == seleccion:
                    txt_presentacion.configure(state="normal")
                    txt_presentacion.delete(0, "end")
                    txt_presentacion.insert(0, item['presentacion'])
                    txt_presentacion.configure(state="readonly")
                    id_catalogo_seleccionado = item['id_catalogo']
                    break

        def on_search(event):
            cargar_catalogo(search_entry.get().strip())

        search_entry.bind("<KeyRelease>", on_search)
        combo_catalogo.configure(command=on_catalogo_seleccionado)

        def toggle_donante():
            if es_donacion.get():
                txt_donante.configure(state="normal")
            else:
                txt_donante.configure(state="disabled")
                txt_donante.delete(0, "end")

        check_donacion.configure(command=toggle_donante)

        # Calendario
        def abrir_calendario():
            self._abrir_calendario_generico_dialog(txt_caducidad, dialog)

        btn_cal.configure(command=abrir_calendario)

        def guardar():
            try:
                if not id_catalogo_seleccionado:
                    messagebox.showerror("Error", "Seleccione un medicamento del catálogo")
                    return

                ubicacion = txt_ubicacion.get().strip()
                stock = txt_stock.get().strip()
                caducidad_str = txt_caducidad.get().strip()
                observaciones = txt_observaciones.get("1.0", "end-1c").strip()
                donacion = es_donacion.get()
                donante = txt_donante.get().strip() if donacion else ""

                try:
                    stock_int = int(stock) if stock else 0
                    if not medicamento_id and stock_int < 1:
                        messagebox.showerror("Error", "El stock inicial debe ser al menos 1 unidad")
                        return
                    if stock_int < 0:
                        messagebox.showerror("Error", "El stock no puede ser negativo")
                        return
                except ValueError:
                    messagebox.showerror("Error", "El stock debe ser un número")
                    return

                caducidad = None
                if caducidad_str:
                    try:
                        caducidad = datetime.strptime(caducidad_str, "%d/%m/%Y").date()
                    except ValueError:
                        messagebox.showerror("Error", "Formato de fecha inválido. Use DD/MM/AAAA")
                        return

                if medicamento_id:
                    success = self.medicamento_model.actualizar(
                        medicamento_id, None, None, ubicacion,
                        caducidad, observaciones, stock_int, donacion, donante
                    )
                    msg = "actualizado"
                else:
                    quien_recibe = "Registro inicial"
                    success = self.medicamento_model.crear_desde_catalogo(
                        self.area_id, id_catalogo_seleccionado, ubicacion,
                        caducidad, stock_int, donacion, donante, observaciones
                    )
                    msg = "agregado"
                    if success and stock_int > 0:
                        medicamentos = self.medicamento_model.obtener_por_area(self.area_id)
                        if medicamentos:
                            nuevo_id = medicamentos[-1]['id_medicamento']
                            self.movimiento_model.registrar(
                                nuevo_id, "entrada", stock_int,
                                quien_recibe, "",
                                f"Registro inicial - {quien_recibe}"
                            )

                if success:
                    messagebox.showinfo("Éxito", f"Medicamento {msg} correctamente")
                    self.cargar_inventario()
                    self.cargar_combo_medicamentos_filtro()
                    self.cargar_combo_ubicaciones()
                    self.cargar_combo_medicamentos_mov()
                    dialog.destroy()
                else:
                    messagebox.showerror("Error", f"No se pudo {msg} el medicamento")
            except Exception as e:
                messagebox.showerror("Error", f"Error al guardar: {str(e)}")

        # Cargar datos si es edición
        if medicamento_id:
            medicamento = self.medicamento_model.obtener_por_id(medicamento_id)
            if medicamento:
                search_entry.insert(0, medicamento['nombre'])
                cargar_catalogo(medicamento['nombre'])
                txt_ubicacion.insert(0, medicamento['ubicacion'])
                txt_stock.insert(0, str(medicamento['stock']))
                if medicamento['caducidad']:
                    txt_caducidad.insert(0, medicamento['caducidad'].strftime("%d/%m/%Y"))
                es_donacion.set(medicamento['es_donacion'])
                if medicamento['es_donacion'] and medicamento['donante']:
                    txt_donante.configure(state="normal")
                    txt_donante.insert(0, medicamento['donante'])
                if medicamento.get('observaciones'):
                    txt_observaciones.insert("1.0", medicamento['observaciones'])
                toggle_donante()
        else:
            cargar_catalogo()

        # Botones
        frame_botones = ctk.CTkFrame(scroll, fg_color="transparent")
        frame_botones.pack(fill="x", pady=15)

        btn_guardar = ctk.CTkButton(frame_botones, text="💾 GUARDAR", height=45,
                                    font=("Segoe UI", 13, "bold"), fg_color=AppTheme.PRIMARY,
                                    hover_color=AppTheme.PRIMARY_DARK, corner_radius=10,
                                    command=guardar)
        btn_guardar.pack(fill="x", pady=2)

        btn_cancelar = ctk.CTkButton(frame_botones, text="❌ CANCELAR", height=40,
                                     font=("Segoe UI", 12), fg_color="#4a5568",
                                     hover_color="#2d3748", text_color=AppTheme.TEXT_ON_PRIMARY,
                                     corner_radius=10, command=dialog.destroy)
        btn_cancelar.pack(fill="x", pady=2)

    # ---------- FORMULARIO MOVIMIENTO RÁPIDO (VENTANA EMERGENTE) ----------
    def abrir_formulario_movimiento_rapido(self, medicamento_id=None):
        dialog = ctk.CTkToplevel(self)
        dialog.title("Registrar Movimiento")
        dialog.geometry("500x650")
        dialog.resizable(False, False)
        dialog.grab_set()

        dialog.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() // 2) - 250
        y = self.winfo_y() + (self.winfo_height() // 2) - 325
        dialog.geometry(f"+{x}+{y}")

        scroll = ctk.CTkScrollableFrame(dialog, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=20, pady=20)

        titulo = ctk.CTkLabel(scroll, text="REGISTRAR MOVIMIENTO",
                              font=("Segoe UI", 18, "bold"), text_color=AppTheme.PRIMARY)
        titulo.pack(pady=(0, 15))

        ctk.CTkLabel(scroll, text="Medicamento", font=("Segoe UI", 12, "bold")).pack(anchor="w", pady=(5, 2))
        combo_medicamento = ctk.CTkComboBox(scroll, values=[], height=38,
                                            font=("Segoe UI", 11), corner_radius=10, state="readonly")
        combo_medicamento.pack(fill="x", pady=(0, 5))

        lbl_stock = ctk.CTkLabel(scroll, text="Stock disponible: -- unidades",
                                 font=("Segoe UI", 11), text_color=AppTheme.PRIMARY)
        lbl_stock.pack(anchor="w", pady=(0, 10))

        ctk.CTkLabel(scroll, text="Tipo de Movimiento", font=("Segoe UI", 12, "bold")).pack(anchor="w", pady=(5, 2))
        tipo_mov = ctk.StringVar(value="entrada")
        frame_tipo = ctk.CTkFrame(scroll, fg_color="transparent")
        frame_tipo.pack(fill="x", pady=5)

        radio_entrada = ctk.CTkRadioButton(frame_tipo, text="ENTRADA", variable=tipo_mov,
                                           value="entrada", font=("Segoe UI", 11, "bold"),
                                           fg_color="#2c7a4d", hover_color="#1e5a38",
                                           text_color=AppTheme.TEXT_PRIMARY)
        radio_entrada.pack(side="left", padx=10)

        radio_salida = ctk.CTkRadioButton(frame_tipo, text="SALIDA", variable=tipo_mov,
                                          value="salida", font=("Segoe UI", 11, "bold"),
                                          fg_color="#e53e3e", hover_color="#c53030",
                                          text_color=AppTheme.TEXT_PRIMARY)
        radio_salida.pack(side="left", padx=10)

        ctk.CTkLabel(scroll, text="Cantidad", font=("Segoe UI", 12, "bold")).pack(anchor="w", pady=(5, 2))
        txt_cantidad = ctk.CTkEntry(scroll, height=38, font=("Segoe UI", 11), corner_radius=10)
        txt_cantidad.pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(scroll, text="Quien Recibe/Retira", font=("Segoe UI", 12, "bold")).pack(anchor="w", pady=(5, 2))
        txt_quien = ctk.CTkEntry(scroll, height=38, font=("Segoe UI", 11), corner_radius=10)
        txt_quien.pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(scroll, text="Motivo", font=("Segoe UI", 12, "bold")).pack(anchor="w", pady=(5, 2))
        txt_motivo = ctk.CTkEntry(scroll, height=38, font=("Segoe UI", 11), corner_radius=10)
        txt_motivo.pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(scroll, text="Fecha", font=("Segoe UI", 12, "bold")).pack(anchor="w", pady=(5, 2))
        frame_fecha = ctk.CTkFrame(scroll, fg_color="transparent")
        frame_fecha.pack(fill="x", pady=(0, 10))
        txt_fecha = ctk.CTkEntry(frame_fecha, placeholder_text="DD/MM/AAAA", height=38,
                                 font=("Segoe UI", 11), corner_radius=10)
        txt_fecha.pack(side="left", fill="x", expand=True, padx=(0, 5))
        txt_fecha.insert(0, date.today().strftime("%d/%m/%Y"))

        btn_cal = ctk.CTkButton(frame_fecha, text="📅", width=50, height=38, font=("Segoe UI", 14),
                                fg_color=AppTheme.PRIMARY, hover_color=AppTheme.PRIMARY_DARK,
                                corner_radius=10)
        btn_cal.pack(side="right")

        def abrir_calendario():
            self._abrir_calendario_generico_dialog(txt_fecha, dialog)

        btn_cal.configure(command=abrir_calendario)

        def actualizar_label_quien():
            if tipo_mov.get() == "entrada":
                txt_quien.configure(placeholder_text="Nombre de quien recibe...")
            else:
                txt_quien.configure(placeholder_text="Nombre de quien retira...")

        def on_tipo_change():
            actualizar_label_quien()
            actualizar_stock()

        def actualizar_stock():
            try:
                seleccion = combo_medicamento.get()
                if seleccion:
                    med_id = int(seleccion.split(" - ")[0])
                    medicamento = self.medicamento_model.obtener_por_id(med_id)
                    if medicamento:
                        stock_actual = medicamento['stock']
                        lbl_stock.configure(text=f"Stock disponible: {stock_actual} unidades")
                        if stock_actual == 0:
                            lbl_stock.configure(text_color="#e53e3e")
                        elif stock_actual < 10:
                            lbl_stock.configure(text_color="#ed8936")
                        else:
                            lbl_stock.configure(text_color=AppTheme.PRIMARY)
            except:
                pass

        radio_entrada.configure(command=on_tipo_change)
        radio_salida.configure(command=on_tipo_change)
        combo_medicamento.configure(command=lambda x: actualizar_stock())
        actualizar_label_quien()

        # Cargar medicamentos
        def cargar_medicamentos():
            medicamentos = self.medicamento_model.obtener_por_area(self.area_id)
            valores = [f"{m['id_medicamento']} - {m['nombre']}" for m in medicamentos]
            combo_medicamento.configure(values=valores)
            if valores:
                combo_medicamento.set(valores[0])
                actualizar_stock()
                if medicamento_id:
                    for v in valores:
                        if str(medicamento_id) in v:
                            combo_medicamento.set(v)
                            actualizar_stock()
                            break

        cargar_medicamentos()

        def registrar():
            try:
                seleccion = combo_medicamento.get()
                if not seleccion:
                    messagebox.showerror("Error", "Seleccione un medicamento")
                    return

                medicamento_id_sel = int(seleccion.split(" - ")[0])
                cantidad = txt_cantidad.get().strip()
                quien = txt_quien.get().strip()
                motivo = txt_motivo.get().strip()
                tipo = tipo_mov.get()
                fecha_str = txt_fecha.get().strip()

                if not cantidad or not quien or not motivo:
                    messagebox.showerror("Error", "Complete todos los campos")
                    return

                try:
                    cantidad_int = int(cantidad)
                    if cantidad_int <= 0:
                        messagebox.showerror("Error", "La cantidad debe ser mayor a 0")
                        return
                except ValueError:
                    messagebox.showerror("Error", "La cantidad debe ser un número")
                    return

                fecha_mov = date.today()
                if fecha_str:
                    try:
                        fecha_mov = datetime.strptime(fecha_str, "%d/%m/%Y").date()
                    except ValueError:
                        messagebox.showerror("Error", "Formato de fecha inválido")
                        return

                medicamento = self.medicamento_model.obtener_por_id(medicamento_id_sel)
                if not medicamento:
                    messagebox.showerror("Error", "Medicamento no encontrado")
                    return

                if tipo == "salida" and medicamento['stock'] < cantidad_int:
                    messagebox.showerror("Error", f"Stock insuficiente.\nStock actual: {medicamento['stock']}")
                    return

                quien_recibe = quien if tipo == "entrada" else ""
                quien_retira = quien if tipo == "salida" else ""

                success = self.movimiento_model.registrar(
                    medicamento_id_sel, tipo, cantidad_int,
                    quien_recibe, quien_retira, motivo, fecha_mov
                )

                if success:
                    nuevo_stock = medicamento['stock']
                    if tipo == "entrada":
                        nuevo_stock += cantidad_int
                    else:
                        nuevo_stock -= cantidad_int
                    self.medicamento_model.actualizar_stock(medicamento_id_sel, nuevo_stock)
                    messagebox.showinfo("Éxito", "Movimiento registrado correctamente")
                    self.cargar_inventario()
                    self.cargar_historial()
                    self.cargar_combo_medicamentos_mov()
                    self.cargar_combo_medicamentos_filtro()
                    dialog.destroy()
                else:
                    messagebox.showerror("Error", "No se pudo registrar el movimiento")
            except Exception as e:
                messagebox.showerror("Error", f"Error: {str(e)}")

        # Botones
        frame_botones = ctk.CTkFrame(scroll, fg_color="transparent")
        frame_botones.pack(fill="x", pady=15)

        btn_registrar = ctk.CTkButton(frame_botones, text="📝 REGISTRAR MOVIMIENTO", height=45,
                                      font=("Segoe UI", 13, "bold"), fg_color=AppTheme.PRIMARY,
                                      hover_color=AppTheme.PRIMARY_DARK, corner_radius=10,
                                      command=registrar)
        btn_registrar.pack(fill="x", pady=2)

        btn_cancelar = ctk.CTkButton(frame_botones, text="❌ CANCELAR", height=40,
                                     font=("Segoe UI", 12), fg_color="#4a5568",
                                     hover_color="#2d3748", text_color=AppTheme.TEXT_ON_PRIMARY,
                                     corner_radius=10, command=dialog.destroy)
        btn_cancelar.pack(fill="x", pady=2)

    # ---------- PESTAÑA MOVIMIENTOS ----------
    def crear_pestana_movimientos(self):
        main_frame = ctk.CTkFrame(self.tab_movimientos, fg_color="transparent")
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)

        # Filtros
        frame_filtros = ctk.CTkFrame(main_frame, fg_color=AppTheme.SURFACE, corner_radius=10)
        frame_filtros.pack(fill="x", pady=(0, 10), padx=5)

        # Fila 1
        row1 = ctk.CTkFrame(frame_filtros, fg_color="transparent")
        row1.pack(fill="x", padx=10, pady=10)

        ctk.CTkLabel(row1, text="Buscar:", font=("Segoe UI", 11)).pack(side="left", padx=5)
        self.txt_buscar_mov = ctk.CTkEntry(row1, placeholder_text="Medicamento, motivo, quien...",
                                           height=35, font=("Segoe UI", 11), corner_radius=10)
        self.txt_buscar_mov.pack(side="left", fill="x", expand=True, padx=5)
        self.txt_buscar_mov.bind("<KeyRelease>", self.buscar_historial)

        # Fila 2
        row2 = ctk.CTkFrame(frame_filtros, fg_color="transparent")
        row2.pack(fill="x", padx=10, pady=(0, 10))

        ctk.CTkLabel(row2, text="Medicamento:", font=("Segoe UI", 11)).pack(side="left", padx=5)
        self.combo_med_filtro = ctk.CTkComboBox(row2, values=[], width=250, height=35,
                                                font=("Segoe UI", 11), corner_radius=10, state="readonly")
        self.combo_med_filtro.pack(side="left", padx=5)
        self.combo_med_filtro.bind("<<ComboboxSelected>>", self.cargar_historial)

        ctk.CTkLabel(row2, text="Tipo:", font=("Segoe UI", 11)).pack(side="left", padx=10)
        self.combo_tipo_filtro = ctk.CTkComboBox(row2, values=["Todos", "Entrada", "Salida"],
                                                 width=100, height=35, corner_radius=10, state="readonly")
        self.combo_tipo_filtro.pack(side="left", padx=5)
        self.combo_tipo_filtro.bind("<<ComboboxSelected>>", self.cargar_historial)
        self.combo_tipo_filtro.set("Todos")

        ctk.CTkLabel(row2, text="Desde:", font=("Segoe UI", 11)).pack(side="left", padx=10)
        self.btn_desde = ctk.CTkButton(row2, text=self.fecha_desde.strftime("%d/%m/%Y"),
                                       width=100, height=35, font=("Segoe UI", 11),
                                       fg_color=AppTheme.PRIMARY, hover_color=AppTheme.PRIMARY_DARK,
                                       corner_radius=10, command=lambda: self.abrir_calendario("desde"))
        self.btn_desde.pack(side="left", padx=5)

        ctk.CTkLabel(row2, text="Hasta:", font=("Segoe UI", 11)).pack(side="left", padx=5)
        self.btn_hasta = ctk.CTkButton(row2, text=self.fecha_hasta.strftime("%d/%m/%Y"),
                                       width=100, height=35, font=("Segoe UI", 11),
                                       fg_color=AppTheme.PRIMARY, hover_color=AppTheme.PRIMARY_DARK,
                                       corner_radius=10, command=lambda: self.abrir_calendario("hasta"))
        self.btn_hasta.pack(side="left", padx=5)

        btn_filtrar = ctk.CTkButton(row2, text="🔍 Filtrar", width=90, height=35,
                                    font=("Segoe UI", 11, "bold"), fg_color=AppTheme.PRIMARY,
                                    hover_color=AppTheme.PRIMARY_DARK, corner_radius=10,
                                    command=self.cargar_historial)
        btn_filtrar.pack(side="left", padx=10)

        btn_exportar = ctk.CTkButton(row2, text="📎 Exportar", width=100, height=35,
                                     font=("Segoe UI", 11), fg_color="#2c7a4d",
                                     hover_color="#1e5a38", corner_radius=10,
                                     command=self.exportar_excel_movimientos)
        btn_exportar.pack(side="left", padx=5)

        # Tabla
        self.crear_tabla_movimientos(main_frame)

    def crear_tabla_movimientos(self, parent):
        self.frame_tabla_movimientos = ctk.CTkFrame(parent, fg_color="transparent")
        self.frame_tabla_movimientos.pack(fill="both", expand=True, padx=5)

        scroll_y = ttk.Scrollbar(self.frame_tabla_movimientos, orient="vertical")
        scroll_x = ttk.Scrollbar(self.frame_tabla_movimientos, orient="horizontal")

        self.tabla_mov = ttk.Treeview(
            self.frame_tabla_movimientos,
            columns=("ID", "Medicamento", "Tipo", "Cantidad", "Quien Recibe", "Quien Retira", "Motivo", "Fecha"),
            show="headings",
            yscrollcommand=scroll_y.set,
            xscrollcommand=scroll_x.set,
            style="Movements.Treeview",
            height=20
        )

        scroll_y.config(command=self.tabla_mov.yview)
        scroll_x.config(command=self.tabla_mov.xview)

        columnas = [
            ("ID", 50, "center"),
            ("Medicamento", 180, "w"),
            ("Tipo", 90, "center"),
            ("Cantidad", 70, "center"),
            ("Quien Recibe", 150, "w"),
            ("Quien Retira", 150, "w"),
            ("Motivo", 220, "w"),
            ("Fecha", 95, "center")
        ]

        for col, width, anchor in columnas:
            self.tabla_mov.heading(col, text=col)
            self.tabla_mov.column(col, width=width, anchor=anchor)

        self.tabla_mov.grid(row=0, column=0, sticky="nsew")
        scroll_y.grid(row=0, column=1, sticky="ns")
        scroll_x.grid(row=1, column=0, sticky="ew")

        self.frame_tabla_movimientos.grid_rowconfigure(0, weight=1)
        self.frame_tabla_movimientos.grid_columnconfigure(0, weight=1)

        # Doble click para editar
        self.tabla_mov.bind("<Double-1>", self.on_movimiento_doble_click)

    def on_movimiento_doble_click(self, event):
        seleccion = self.tabla_mov.selection()
        if seleccion:
            valores = self.tabla_mov.item(seleccion[0])['values']
            if valores and valores[0]:
                self.abrir_editar_movimiento(valores[0])

    def abrir_editar_movimiento(self, movimiento_id):
        # Similar al formulario de movimiento pero con datos cargados
        movimiento = self.movimiento_model.obtener_por_id(movimiento_id, self.area_id)
        if not movimiento:
            return

        dialog = ctk.CTkToplevel(self)
        dialog.title("Editar Movimiento")
        dialog.geometry("500x650")
        dialog.resizable(False, False)
        dialog.grab_set()

        dialog.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() // 2) - 250
        y = self.winfo_y() + (self.winfo_height() // 2) - 325
        dialog.geometry(f"+{x}+{y}")

        scroll = ctk.CTkScrollableFrame(dialog, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=20, pady=20)

        titulo = ctk.CTkLabel(scroll, text="EDITAR MOVIMIENTO",
                              font=("Segoe UI", 18, "bold"), text_color=AppTheme.PRIMARY)
        titulo.pack(pady=(0, 15))

        ctk.CTkLabel(scroll, text="Medicamento", font=("Segoe UI", 12, "bold")).pack(anchor="w", pady=(5, 2))
        combo_medicamento = ctk.CTkComboBox(scroll, values=[], height=38,
                                            font=("Segoe UI", 11), corner_radius=10, state="readonly")
        combo_medicamento.pack(fill="x", pady=(0, 5))

        lbl_stock = ctk.CTkLabel(scroll, text="Stock disponible: -- unidades",
                                 font=("Segoe UI", 11), text_color=AppTheme.PRIMARY)
        lbl_stock.pack(anchor="w", pady=(0, 10))

        ctk.CTkLabel(scroll, text="Tipo de Movimiento", font=("Segoe UI", 12, "bold")).pack(anchor="w", pady=(5, 2))
        tipo_mov = ctk.StringVar(value=movimiento['tipo'])
        frame_tipo = ctk.CTkFrame(scroll, fg_color="transparent")
        frame_tipo.pack(fill="x", pady=5)

        radio_entrada = ctk.CTkRadioButton(frame_tipo, text="ENTRADA", variable=tipo_mov,
                                           value="entrada", font=("Segoe UI", 11, "bold"),
                                           fg_color="#2c7a4d", hover_color="#1e5a38",
                                           text_color=AppTheme.TEXT_PRIMARY)
        radio_entrada.pack(side="left", padx=10)

        radio_salida = ctk.CTkRadioButton(frame_tipo, text="SALIDA", variable=tipo_mov,
                                          value="salida", font=("Segoe UI", 11, "bold"),
                                          fg_color="#e53e3e", hover_color="#c53030",
                                          text_color=AppTheme.TEXT_PRIMARY)
        radio_salida.pack(side="left", padx=10)

        ctk.CTkLabel(scroll, text="Cantidad", font=("Segoe UI", 12, "bold")).pack(anchor="w", pady=(5, 2))
        txt_cantidad = ctk.CTkEntry(scroll, height=38, font=("Segoe UI", 11), corner_radius=10)
        txt_cantidad.pack(fill="x", pady=(0, 10))
        txt_cantidad.insert(0, str(movimiento['cantidad']))

        ctk.CTkLabel(scroll, text="Quien Recibe/Retira", font=("Segoe UI", 12, "bold")).pack(anchor="w", pady=(5, 2))
        txt_quien = ctk.CTkEntry(scroll, height=38, font=("Segoe UI", 11), corner_radius=10)
        txt_quien.pack(fill="x", pady=(0, 10))
        if movimiento['tipo'] == "entrada":
            txt_quien.insert(0, movimiento['quien_recibe'] or "")
        else:
            txt_quien.insert(0, movimiento['quien_retira'] or "")

        ctk.CTkLabel(scroll, text="Motivo", font=("Segoe UI", 12, "bold")).pack(anchor="w", pady=(5, 2))
        txt_motivo = ctk.CTkEntry(scroll, height=38, font=("Segoe UI", 11), corner_radius=10)
        txt_motivo.pack(fill="x", pady=(0, 10))
        txt_motivo.insert(0, movimiento['motivo'] or "")

        ctk.CTkLabel(scroll, text="Fecha", font=("Segoe UI", 12, "bold")).pack(anchor="w", pady=(5, 2))
        frame_fecha = ctk.CTkFrame(scroll, fg_color="transparent")
        frame_fecha.pack(fill="x", pady=(0, 10))
        txt_fecha = ctk.CTkEntry(frame_fecha, placeholder_text="DD/MM/AAAA", height=38,
                                 font=("Segoe UI", 11), corner_radius=10)
        txt_fecha.pack(side="left", fill="x", expand=True, padx=(0, 5))
        txt_fecha.insert(0, movimiento['fecha'].strftime("%d/%m/%Y") if movimiento['fecha'] else "")

        btn_cal = ctk.CTkButton(frame_fecha, text="📅", width=50, height=38, font=("Segoe UI", 14),
                                fg_color=AppTheme.PRIMARY, hover_color=AppTheme.PRIMARY_DARK,
                                corner_radius=10)
        btn_cal.pack(side="right")

        def abrir_calendario():
            self._abrir_calendario_generico_dialog(txt_fecha, dialog)

        btn_cal.configure(command=abrir_calendario)

        def actualizar_stock():
            try:
                seleccion = combo_medicamento.get()
                if seleccion:
                    med_id = int(seleccion.split(" - ")[0])
                    medicamento = self.medicamento_model.obtener_por_id(med_id)
                    if medicamento:
                        stock_actual = medicamento['stock']
                        lbl_stock.configure(text=f"Stock disponible: {stock_actual} unidades")
            except:
                pass

        combo_medicamento.configure(command=lambda x: actualizar_stock())

        # Cargar medicamentos y seleccionar el actual
        medicamentos = self.medicamento_model.obtener_por_area(self.area_id)
        valores = [f"{m['id_medicamento']} - {m['nombre']}" for m in medicamentos]
        combo_medicamento.configure(values=valores)
        for v in valores:
            if str(movimiento['id_medicamento']) in v.split(" - ")[0]:
                combo_medicamento.set(v)
                break
        actualizar_stock()

        def actualizar():
            try:
                seleccion = combo_medicamento.get()
                if not seleccion:
                    messagebox.showerror("Error", "Seleccione un medicamento")
                    return

                nuevo_medicamento_id = int(seleccion.split(" - ")[0])
                nueva_cantidad = int(txt_cantidad.get().strip())
                nuevo_tipo = tipo_mov.get()
                nuevo_quien = txt_quien.get().strip()
                nuevo_motivo = txt_motivo.get().strip()
                nueva_fecha_str = txt_fecha.get().strip()

                if nueva_cantidad <= 0:
                    messagebox.showerror("Error", "La cantidad debe ser mayor a 0")
                    return

                if not nuevo_motivo:
                    messagebox.showerror("Error", "El motivo es obligatorio")
                    return

                nueva_fecha = date.today()
                if nueva_fecha_str:
                    try:
                        nueva_fecha = datetime.strptime(nueva_fecha_str, "%d/%m/%Y").date()
                    except ValueError:
                        messagebox.showerror("Error", "Formato de fecha inválido")
                        return

                # Calcular nuevo stock
                medicamento_actual = self.medicamento_model.obtener_por_id(movimiento['id_medicamento'])
                nuevo_medicamento = self.medicamento_model.obtener_por_id(nuevo_medicamento_id)

                # Revertir efecto del movimiento original
                stock_temporal = medicamento_actual['stock']
                if movimiento['tipo'] == "entrada":
                    stock_temporal -= movimiento['cantidad']
                else:
                    stock_temporal += movimiento['cantidad']

                # Aplicar nuevo movimiento
                if nuevo_tipo == "entrada":
                    stock_final = stock_temporal + nueva_cantidad
                else:
                    if stock_temporal < nueva_cantidad:
                        messagebox.showerror("Error", f"Stock insuficiente. Stock actual: {stock_temporal}")
                        return
                    stock_final = stock_temporal - nueva_cantidad

                quien_recibe = nuevo_quien if nuevo_tipo == "entrada" else ""
                quien_retira = nuevo_quien if nuevo_tipo == "salida" else ""

                tabla_movimientos = "movimientos_farmacia" if self.area_id == 1 else "movimientos_odontologia"
                query = f"""
                    UPDATE {tabla_movimientos}
                    SET tipo = ?, cantidad = ?, quien_recibe = ?, quien_retira = ?, motivo = ?, fecha = ?, id_inventario = ?
                    WHERE id_movimiento = ?
                """
                params = (nuevo_tipo, nueva_cantidad, quien_recibe, quien_retira,
                          nuevo_motivo, nueva_fecha, nuevo_medicamento_id, movimiento_id)
                success = Database.execute_non_query(query, params)

                if success:
                    # Actualizar stocks
                    self.medicamento_model.actualizar_stock(movimiento['id_medicamento'], stock_temporal)
                    if nuevo_medicamento_id != movimiento['id_medicamento']:
                        self.medicamento_model.actualizar_stock(nuevo_medicamento_id, stock_final)
                    else:
                        self.medicamento_model.actualizar_stock(nuevo_medicamento_id, stock_final)

                    messagebox.showinfo("Éxito", "Movimiento actualizado correctamente")
                    self.cargar_historial()
                    self.cargar_inventario()
                    self.cargar_combo_medicamentos_filtro()
                    self.cargar_combo_medicamentos_mov()
                    dialog.destroy()
                else:
                    messagebox.showerror("Error", "No se pudo actualizar el movimiento")
            except ValueError:
                messagebox.showerror("Error", "La cantidad debe ser un número")
            except Exception as e:
                messagebox.showerror("Error", f"Error: {str(e)}")

        # Botones
        frame_botones = ctk.CTkFrame(scroll, fg_color="transparent")
        frame_botones.pack(fill="x", pady=15)

        btn_actualizar = ctk.CTkButton(frame_botones, text="💾 ACTUALIZAR", height=45,
                                       font=("Segoe UI", 13, "bold"), fg_color=AppTheme.PRIMARY,
                                       hover_color=AppTheme.PRIMARY_DARK, corner_radius=10,
                                       command=actualizar)
        btn_actualizar.pack(fill="x", pady=2)

        btn_cancelar = ctk.CTkButton(frame_botones, text="❌ CANCELAR", height=40,
                                     font=("Segoe UI", 12), fg_color="#4a5568",
                                     hover_color="#2d3748", text_color=AppTheme.TEXT_ON_PRIMARY,
                                     corner_radius=10, command=dialog.destroy)
        btn_cancelar.pack(fill="x", pady=2)

    def cargar_combo_medicamentos_mov(self):
        medicamentos = self.medicamento_model.obtener_por_area(self.area_id)
        # No es necesario cargar aquí, se usa en los diálogos

    def cargar_combo_medicamentos_filtro(self):
        try:
            medicamentos = self.medicamento_model.obtener_por_area(self.area_id)
            valores = ["-- TODOS --"] + [f"{m['id_medicamento']} - {m['nombre']}" for m in medicamentos]
            self.combo_med_filtro.configure(values=valores)
            if valores:
                self.combo_med_filtro.set(valores[0])
        except Exception as e:
            self.debug_print(f"Error al cargar combo filtro: {e}")

    def cargar_historial(self, event=None):
        try:
            for item in self.tabla_mov.get_children():
                self.tabla_mov.delete(item)

            busqueda = self.txt_buscar_mov.get().strip().lower()
            seleccion = self.combo_med_filtro.get()
            if not seleccion:
                return

            fecha_desde = self.fecha_desde
            fecha_hasta = self.fecha_hasta
            tipo_filtro = self.combo_tipo_filtro.get() if self.combo_tipo_filtro.get() != "Todos" else None

            movimientos = self.movimiento_model.obtener_todos_por_area(
                self.area_id, fecha_desde, fecha_hasta, tipo_filtro
            )

            if not movimientos:
                self.tabla_mov.insert("", "end",
                                      values=("", "", "", "", "", "", "No hay movimientos", ""))
                return

            for mov in reversed(movimientos):
                nombre_med = mov.get('medicamento_nombre', '')
                motivo = mov.get('motivo', '')
                quien_recibe = mov.get('quien_recibe', '---')
                quien_retira = mov.get('quien_retira', '---')

                if seleccion != "-- TODOS --" and seleccion:
                    med_id_filtro = int(seleccion.split(" - ")[0])
                    if mov.get('id_medicamento') != med_id_filtro:
                        continue

                if busqueda:
                    if (busqueda not in motivo.lower() and
                            busqueda not in quien_recibe.lower() and
                            busqueda not in quien_retira.lower() and
                            busqueda not in nombre_med.lower()):
                        continue

                fecha = mov['fecha'].strftime("%d/%m/%Y") if isinstance(mov['fecha'], date) else str(mov['fecha'])
                tipo = "ENTRADA" if mov['tipo'] == "entrada" else "SALIDA"

                self.tabla_mov.insert("", "end", values=(
                    mov['id_movimiento'], nombre_med, tipo, mov['cantidad'],
                    quien_recibe, quien_retira, motivo, fecha
                ))
        except Exception as e:
            self.debug_print(f"Error cargando historial: {e}")
            messagebox.showerror("Error", f"Error cargando historial: {str(e)}")

    def buscar_historial(self, event):
        self.cargar_historial()

    def abrir_calendario(self, tipo):
        # Calendario para fechas de filtro
        dialog = ctk.CTkToplevel(self)
        dialog.title("Seleccionar fecha")
        dialog.geometry("350x450")
        dialog.resizable(False, False)
        dialog.grab_set()

        dialog.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() // 2) - 175
        y = self.winfo_y() + (self.winfo_height() // 2) - 225
        dialog.geometry(f"+{x}+{y}")

        frame_cal = ctk.CTkFrame(dialog, fg_color=AppTheme.SURFACE, corner_radius=20)
        frame_cal.pack(fill="both", expand=True, padx=15, pady=15)

        cal_year = datetime.now().year
        cal_month = datetime.now().month

        nav_frame = ctk.CTkFrame(frame_cal, fg_color="transparent")
        nav_frame.pack(fill="x", pady=10)

        def cambiar_mes(delta):
            nonlocal cal_year, cal_month
            cal_month += delta
            if cal_month > 12:
                cal_month = 1
                cal_year += 1
            elif cal_month < 1:
                cal_month = 12
                cal_year -= 1
            lbl_mes_anio.configure(text=f"{self.obtener_nombre_mes(cal_month)} {cal_year}")
            actualizar_calendario()

        def obtener_nombre_mes(mes):
            meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                     "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
            return meses[mes - 1]

        btn_prev = ctk.CTkButton(nav_frame, text="<", width=35, height=35, fg_color=AppTheme.PRIMARY,
                                 hover_color=AppTheme.PRIMARY_DARK,
                                 command=lambda: cambiar_mes(-1))
        btn_prev.pack(side="left", padx=5)

        lbl_mes_anio = ctk.CTkLabel(nav_frame, text=obtener_nombre_mes(cal_month) + " " + str(cal_year),
                                    font=("Segoe UI", 16, "bold"), text_color=AppTheme.PRIMARY)
        lbl_mes_anio.pack(side="left", expand=True)

        btn_next = ctk.CTkButton(nav_frame, text=">", width=35, height=35, fg_color=AppTheme.PRIMARY,
                                 hover_color=AppTheme.PRIMARY_DARK,
                                 command=lambda: cambiar_mes(1))
        btn_next.pack(side="right", padx=5)

        dias_frame = ctk.CTkFrame(frame_cal, fg_color="transparent")
        dias_frame.pack(fill="both", expand=True, pady=5)

        # Días de semana
        dias_semana = ctk.CTkFrame(frame_cal, fg_color="transparent")
        dias_semana.pack(fill="x", pady=10)
        for dia in ["Lun", "Mar", "Mie", "Jue", "Vie", "Sab", "Dom"]:
            label = ctk.CTkLabel(dias_semana, text=dia, width=45, font=("Segoe UI", 11, "bold"),
                                 text_color=AppTheme.PRIMARY)
            label.pack(side="left", expand=True, fill="x")

        def actualizar_calendario():
            for widget in dias_frame.winfo_children():
                widget.destroy()

            cal = calendar.monthcalendar(cal_year, cal_month)
            hoy = date.today()

            for semana in cal:
                frame_semana = ctk.CTkFrame(dias_frame, fg_color="transparent")
                frame_semana.pack(fill="x", pady=2)

                for dia in semana:
                    if dia == 0:
                        btn = ctk.CTkButton(frame_semana, text="", width=45, height=38,
                                            fg_color="transparent", state="disabled")
                    else:
                        fecha = date(cal_year, cal_month, dia)
                        if fecha == hoy:
                            bg_color = "#F39C12"
                            text_color = "#FFFFFF"
                        else:
                            bg_color = "transparent"
                            text_color = AppTheme.TEXT_PRIMARY

                        btn = ctk.CTkButton(
                            frame_semana, text=str(dia), width=45, height=38,
                            font=("Segoe UI", 11), fg_color=bg_color,
                            hover_color=AppTheme.PRIMARY_LIGHT, text_color=text_color,
                            corner_radius=8,
                            command=lambda d=dia, f=fecha: seleccionar_fecha(f)
                        )
                    btn.pack(side="left", expand=True, fill="x", padx=2)

        def seleccionar_fecha(fecha):
            if tipo == "desde":
                self.fecha_desde = fecha
                self.btn_desde.configure(text=fecha.strftime("%d/%m/%Y"))
            else:
                self.fecha_hasta = fecha
                self.btn_hasta.configure(text=fecha.strftime("%d/%m/%Y"))
            self.cargar_historial()
            dialog.destroy()

        actualizar_calendario()

        btn_aceptar = ctk.CTkButton(frame_cal, text="Aceptar", width=100, height=35,
                                    font=("Segoe UI", 11, "bold"), fg_color=AppTheme.PRIMARY,
                                    hover_color=AppTheme.PRIMARY_DARK, command=dialog.destroy)
        btn_aceptar.pack(pady=10)

    def _abrir_calendario_generico_dialog(self, entry_widget, parent_dialog):
        """Calendario para campos de fecha en diálogos"""
        dialog = ctk.CTkToplevel(parent_dialog)
        dialog.title("Seleccionar fecha")
        dialog.geometry("350x450")
        dialog.resizable(False, False)
        dialog.grab_set()

        dialog.update_idletasks()
        x = parent_dialog.winfo_x() + (parent_dialog.winfo_width() // 2) - 175
        y = parent_dialog.winfo_y() + (parent_dialog.winfo_height() // 2) - 225
        dialog.geometry(f"+{x}+{y}")

        frame_cal = ctk.CTkFrame(dialog, fg_color=AppTheme.SURFACE, corner_radius=20)
        frame_cal.pack(fill="both", expand=True, padx=15, pady=15)

        cal_year = datetime.now().year
        cal_month = datetime.now().month

        nav_frame = ctk.CTkFrame(frame_cal, fg_color="transparent")
        nav_frame.pack(fill="x", pady=10)

        def cambiar_mes(delta):
            nonlocal cal_year, cal_month
            cal_month += delta
            if cal_month > 12:
                cal_month = 1
                cal_year += 1
            elif cal_month < 1:
                cal_month = 12
                cal_year -= 1
            lbl_mes_anio.configure(text=f"{self.obtener_nombre_mes(cal_month)} {cal_year}")
            actualizar_calendario()

        def obtener_nombre_mes(mes):
            meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                     "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
            return meses[mes - 1]

        btn_prev = ctk.CTkButton(nav_frame, text="<", width=35, height=35, fg_color=AppTheme.PRIMARY,
                                 hover_color=AppTheme.PRIMARY_DARK,
                                 command=lambda: cambiar_mes(-1))
        btn_prev.pack(side="left", padx=5)

        lbl_mes_anio = ctk.CTkLabel(nav_frame, text=obtener_nombre_mes(cal_month) + " " + str(cal_year),
                                    font=("Segoe UI", 16, "bold"), text_color=AppTheme.PRIMARY)
        lbl_mes_anio.pack(side="left", expand=True)

        btn_next = ctk.CTkButton(nav_frame, text=">", width=35, height=35, fg_color=AppTheme.PRIMARY,
                                 hover_color=AppTheme.PRIMARY_DARK,
                                 command=lambda: cambiar_mes(1))
        btn_next.pack(side="right", padx=5)

        dias_frame = ctk.CTkFrame(frame_cal, fg_color="transparent")
        dias_frame.pack(fill="both", expand=True, pady=5)

        # Días de semana
        dias_semana = ctk.CTkFrame(frame_cal, fg_color="transparent")
        dias_semana.pack(fill="x", pady=10)
        for dia in ["Lun", "Mar", "Mie", "Jue", "Vie", "Sab", "Dom"]:
            label = ctk.CTkLabel(dias_semana, text=dia, width=45, font=("Segoe UI", 11, "bold"),
                                 text_color=AppTheme.PRIMARY)
            label.pack(side="left", expand=True, fill="x")

        def actualizar_calendario():
            for widget in dias_frame.winfo_children():
                widget.destroy()

            cal = calendar.monthcalendar(cal_year, cal_month)
            hoy = date.today()

            for semana in cal:
                frame_semana = ctk.CTkFrame(dias_frame, fg_color="transparent")
                frame_semana.pack(fill="x", pady=2)

                for dia in semana:
                    if dia == 0:
                        btn = ctk.CTkButton(frame_semana, text="", width=45, height=38,
                                            fg_color="transparent", state="disabled")
                    else:
                        fecha = date(cal_year, cal_month, dia)
                        if fecha == hoy:
                            bg_color = "#F39C12"
                            text_color = "#FFFFFF"
                        else:
                            bg_color = "transparent"
                            text_color = AppTheme.TEXT_PRIMARY

                        btn = ctk.CTkButton(
                            frame_semana, text=str(dia), width=45, height=38,
                            font=("Segoe UI", 11), fg_color=bg_color,
                            hover_color=AppTheme.PRIMARY_LIGHT, text_color=text_color,
                            corner_radius=8,
                            command=lambda d=dia, f=fecha: seleccionar_fecha(f)
                        )
                    btn.pack(side="left", expand=True, fill="x", padx=2)

        def seleccionar_fecha(fecha):
            entry_widget.delete(0, "end")
            entry_widget.insert(0, fecha.strftime("%d/%m/%Y"))
            dialog.destroy()

        actualizar_calendario()

        btn_aceptar = ctk.CTkButton(frame_cal, text="Aceptar", width=100, height=35,
                                    font=("Segoe UI", 11, "bold"), fg_color=AppTheme.PRIMARY,
                                    hover_color=AppTheme.PRIMARY_DARK, command=dialog.destroy)
        btn_aceptar.pack(pady=10)

    # ---------- EXPORTAR EXCEL ----------
    def exportar_excel_inventario(self):
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror("Error", "Instale openpyxl: pip install openpyxl")
            return
        try:
            medicamentos = self.medicamento_model.obtener_por_area(self.area_id, busqueda="")
            if not medicamentos:
                messagebox.showwarning("Sin datos", "No hay medicamentos para exportar.")
                return
            fecha_hora = datetime.now().strftime("%Y%m%d_%H%M%S")
            nombre_por_defecto = f"MediKinv_{self.area_nombre}_Inventario_{fecha_hora}.xlsx"
            archivo = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Archivos Excel", "*.xlsx")],
                initialfile=nombre_por_defecto,
                title="Guardar exportación de inventario"
            )
            if not archivo:
                return
            wb = Workbook()
            ws = wb.active
            ws.title = "Inventario"
            titulo_font = Font(name='Segoe UI', size=16, bold=True, color='FFFFFF')
            titulo_fill = PatternFill(start_color='2C7A4D', end_color='2C7A4D', fill_type='solid')
            header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4A5568', end_color='4A5568', fill_type='solid')
            center_alignment = Alignment(horizontal='center', vertical='center')
            left_alignment = Alignment(horizontal='left', vertical='center')
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))
            ws.merge_cells('A1:I1')
            ws['A1'].value = f"MediKinv - {self.area_nombre} - Inventario"
            ws['A1'].font = titulo_font
            ws['A1'].fill = titulo_fill
            ws['A1'].alignment = center_alignment
            ws.row_dimensions[2].height = 10
            headers = ["ID", "Medicamento", "Presentación", "Ubicación", "Stock", "Caducidad", "Donación", "Donante",
                       "Observaciones"]
            for col_idx, header in enumerate(headers, start=1):
                celda = ws.cell(row=3, column=col_idx, value=header)
                celda.font = header_font
                celda.fill = header_fill
                celda.alignment = center_alignment
                celda.border = thin_border
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                ws.column_dimensions[col].width = 15
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['I'].width = 30
            fila_actual = 4
            for med in medicamentos:
                caducidad = med['caducidad'].strftime("%d/%m/%Y") if med['caducidad'] else "---"
                donacion = "Sí" if med['es_donacion'] else "No"
                ws.cell(row=fila_actual, column=1, value=med['id_medicamento']).alignment = center_alignment
                ws.cell(row=fila_actual, column=2, value=med['nombre']).alignment = left_alignment
                ws.cell(row=fila_actual, column=3, value=med['presentacion']).alignment = left_alignment
                ws.cell(row=fila_actual, column=4, value=med['ubicacion'] or '').alignment = left_alignment
                ws.cell(row=fila_actual, column=5, value=med['stock']).alignment = center_alignment
                ws.cell(row=fila_actual, column=6, value=caducidad).alignment = center_alignment
                ws.cell(row=fila_actual, column=7, value=donacion).alignment = center_alignment
                ws.cell(row=fila_actual, column=8, value=med['donante'] or '').alignment = left_alignment
                ws.cell(row=fila_actual, column=9, value=med.get('observaciones', '') or '').alignment = left_alignment
                for col in range(1, 10):
                    ws.cell(row=fila_actual, column=col).border = thin_border
                fila_actual += 1
            wb.save(archivo)
            messagebox.showinfo("Exportación exitosa", f"Inventario exportado a:\n{archivo}")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo exportar inventario:\n{str(e)}")

    def exportar_excel_movimientos(self):
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror("Error", "Instale openpyxl: pip install openpyxl")
            return
        try:
            movimientos_data = []
            for child in self.tabla_mov.get_children():
                valores = self.tabla_mov.item(child)['values']
                if valores and valores[0] != "" and valores[0] != "No hay movimientos":
                    movimientos_data.append({
                        'id': valores[0],
                        'medicamento': valores[1],
                        'tipo': valores[2],
                        'cantidad': valores[3],
                        'quien_recibe': valores[4],
                        'quien_retira': valores[5],
                        'motivo': valores[6],
                        'fecha': valores[7]
                    })
            if not movimientos_data:
                messagebox.showwarning("Sin datos", "No hay movimientos para exportar.")
                return
            fecha_hora = datetime.now().strftime("%Y%m%d_%H%M%S")
            nombre_por_defecto = f"MediKinv_{self.area_nombre}_Movimientos_{fecha_hora}.xlsx"
            archivo = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Archivos Excel", "*.xlsx")],
                initialfile=nombre_por_defecto,
                title="Guardar exportación de movimientos"
            )
            if not archivo:
                return
            wb = Workbook()
            ws = wb.active
            ws.title = "Movimientos"
            titulo_font = Font(name='Segoe UI', size=16, bold=True, color='FFFFFF')
            titulo_fill = PatternFill(start_color='2C7A4D', end_color='2C7A4D', fill_type='solid')
            header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4A5568', end_color='4A5568', fill_type='solid')
            center_alignment = Alignment(horizontal='center', vertical='center')
            left_alignment = Alignment(horizontal='left', vertical='center')
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))
            ws.merge_cells('A1:H1')
            ws['A1'].value = f"MediKinv - {self.area_nombre} - Movimientos"
            ws['A1'].font = titulo_font
            ws['A1'].fill = titulo_fill
            ws['A1'].alignment = center_alignment
            ws.row_dimensions[2].height = 10
            headers = ["ID", "Medicamento", "Tipo", "Cantidad", "Quien Recibe", "Quien Retira", "Motivo", "Fecha"]
            for col_idx, header in enumerate(headers, start=1):
                celda = ws.cell(row=3, column=col_idx, value=header)
                celda.font = header_font
                celda.fill = header_fill
                celda.alignment = center_alignment
                celda.border = thin_border
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 20
            ws.column_dimensions['G'].width = 35
            ws.column_dimensions['H'].width = 15
            fila_actual = 4
            for mov in movimientos_data:
                ws.cell(row=fila_actual, column=1, value=mov['id']).alignment = center_alignment
                ws.cell(row=fila_actual, column=2, value=mov['medicamento']).alignment = left_alignment
                ws.cell(row=fila_actual, column=3, value=mov['tipo']).alignment = center_alignment
                ws.cell(row=fila_actual, column=4, value=mov['cantidad']).alignment = center_alignment
                ws.cell(row=fila_actual, column=5, value=mov['quien_recibe']).alignment = left_alignment
                ws.cell(row=fila_actual, column=6, value=mov['quien_retira']).alignment = left_alignment
                ws.cell(row=fila_actual, column=7, value=mov['motivo']).alignment = left_alignment
                ws.cell(row=fila_actual, column=8, value=mov['fecha']).alignment = center_alignment
                for col in range(1, 9):
                    ws.cell(row=fila_actual, column=col).border = thin_border
                fila_actual += 1
            wb.save(archivo)
            messagebox.showinfo("Exportación exitosa", f"Movimientos exportados a:\n{archivo}")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo exportar movimientos:\n{str(e)}")

    def obtener_nombre_mes(self, mes):
        meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                 "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
        return meses[mes - 1]


if __name__ == "__main__":
    app = SistemaView(None, 1, "Farmacia")
    app.mainloop()